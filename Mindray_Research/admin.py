
from django.contrib import admin
from django.contrib.auth.admin import UserAdmin
from django.contrib import admin
from django.utils.html import format_html
from django.db import transaction
from Mindray_Research.models import *
from django.utils.translation import gettext_lazy
from django import forms
from django.contrib.admin.widgets import AdminDateWidget
from django.contrib.admin import SimpleListFilter
from django.db.models import Q


from django.forms import DateInput
from datetime import datetime
import nested_admin

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse
from django.utils import timezone
from datetime import datetime
from urllib.parse import quote
from openpyxl.styles import Border, Side
from django.contrib.admin import SimpleListFilter
from django.utils.safestring import mark_safe

from django.db.models import Count, Sum, Q, Avg
from django.utils import timezone
from urllib.parse import quote
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict



ALLOWED_SALESMEN = ['ybb', 'fzj', 'zy', 'cy', 'gsj', 'xxh']

class YearMonthDateWidget(DateInput):
    def __init__(self, attrs=None):
        default_attrs = {
            'type': 'date',
            'class': 'vDateField',
        }
        if attrs:
            default_attrs.update(attrs)
        super().__init__(attrs=default_attrs)
    
    def format_value(self, value):
        """确保值以正确格式显示在HTML date input中"""
        if value is None:
            return ''
        
        # 如果是字符串，尝试解析为日期
        if isinstance(value, str):
            try:
                # 尝试不同的日期格式
                for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%m/%d/%Y', '%d/%m/%Y']:
                    try:
                        value = datetime.strptime(value, fmt).date()
                        break
                    except ValueError:
                        continue
                else:
                    return value  # 如果都解析失败，返回原值
            except:
                return value
        
        # 如果是datetime对象，转换为date
        if hasattr(value, 'date'):
            value = value.date()
        
        # 如果是date对象，格式化为HTML date input需要的格式
        if hasattr(value, 'strftime'):
            return value.strftime('%Y-%m-%d')
        
        return value

    class Media:
        css = {
            'all': (
                'admin/css/widgets.css',
            )
        }
        js = (
            'admin/js/core.js',
            'admin/js/admin/DateTimeShortcuts.js',
        )


class GlobalAdmin(admin.ModelAdmin):
    def delete_queryset(self,request, queryset):
        print('im in global delete_queryset')
        queryset.update(is_active=False)

    def delete_model(self, request, obj):
        print('im in global delete_model')
        obj.is_active = False 
        obj.save()

    def get_queryset(self, request):
        print('im in global get_queryset')
        return super().get_queryset(request).filter(is_active=True)


    class Media:
        css = {
            'all': ('admin/css/custom_fieldsets.css',)
        }


class SalesmanFilter(SimpleListFilter):
    title = '其田负责人' 
    parameter_name = 'userinfo'

    def lookups(self, request, model_admin):

        salesmans = UserInfoMindray.objects.filter(Q(username__in= ALLOWED_SALESMEN))
        # print([(salesman.id, salesman.chinesename) for salesman in salesmans])
        return [(salesman.id, salesman.chinesename) for salesman in salesmans]
    
    def queryset(self, request, queryset):
        if self.value():
        # 筛选条件有值时, 查询对应的 node 的文章
            return queryset.filter(qitian_manager__id=self.value())
        else:
        # 筛选条件没有值时，全部的时候是没有值的
            return queryset


# 1. 创建医院负责人筛选器
class HospitalManagerFilter(SimpleListFilter):
    title = '医院负责人'
    parameter_name = 'hospital_manager'

    def lookups(self, request, model_admin):
        # 获取所有医院调研中的负责人
        managers = UserInfoMindray.objects.filter(
            Q(username__in=ALLOWED_SALESMEN),
            qitian_manager__isnull=False
        ).distinct()
        return [(manager.id, manager.chinesename) for manager in managers]
    
    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(hospital_survey__qitian_manager__id=self.value())
        return queryset

# 2. 创建装机年份筛选器（修正None值问题）
class InstallationYearFilter(SimpleListFilter):
    title = '装机年份'
    parameter_name = 'installation_year'

    def lookups(self, request, model_admin):
        # 获取所有装机年份，包括None值
        years = MindrayInstrumentSurvey.objects.filter(
            is_active=True
        ).values_list('installation_year', flat=True).distinct().order_by('-installation_year')
        
        lookups = []
        for year in years:
            if year:
                lookups.append((year, year))
            else:
                lookups.append(('None', '未填写'))
        return lookups
    
    def queryset(self, request, queryset):
        if self.value():
            if self.value() == 'None':
                return queryset.filter(installation_year__isnull=True)
            else:
                return queryset.filter(installation_year=self.value())
        return queryset
    
 
@admin.register(UserInfoMindray)  
class UserInfoMindrayAdmin(UserAdmin):  
    list_display = ('username','chinesename','first_name','last_name','email','is_staff','is_superuser','date_joined','last_login')
    search_fields = ['username', 'chinesename', 'first_name', 'last_name']
    
    fieldsets = (
        (None, {u'fields': ('username', 'password')}),
        (gettext_lazy('基本信息'), {'fields': (
         'chinesename', 'first_name', 'last_name', 'email',)}),
        (gettext_lazy('权限信息'), {'fields': ('is_superuser', 'is_staff',  'groups', 'user_permissions','is_active')}),
        (gettext_lazy('日期信息'), {'fields': ('last_login', 'date_joined')}),
    )
    
    def get_search_results(self, request, queryset, search_term):
        queryset, use_distinct = super().get_search_results(request, queryset, search_term)
        
        if 'autocomplete' in request.path:
            queryset = queryset.filter(
                is_active=True,
                username__in=ALLOWED_SALESMEN
            ).order_by('chinesename')
        
        return queryset, use_distinct
    
    def has_change_permission(self, request, obj=None):
        return False
    def has_delete_permission(self, request, obj=None):
        return False  # 全局禁用删除按钮
    

@admin.register(Hospital)  
class HospitalAdmin(GlobalAdmin):   
    search_fields=['hospitalname']
    exclude = ('id','createtime','updatetime','is_active')

    def get_search_results(self, request, queryset, search_term):
        queryset,use_distinct = super().get_search_results(request, queryset, search_term)
        if 'autocomplete' in request.path:
            queryset=queryset.filter(is_active=True).order_by('id')
        return queryset,use_distinct 
   
    def has_change_permission(self, request, obj=None):
        return False
    def has_delete_permission(self, request, obj=None):
        return False  # 全局禁用删除按钮
    

@admin.register(Brand)  
class BrandAdmin(GlobalAdmin):   
    search_fields=['brand']
    exclude = ('id','createtime','updatetime','is_active')
    
    def get_search_results(self, request, queryset, search_term):
        queryset,use_distinct = super().get_search_results(request, queryset, search_term)
        if 'autocomplete' in request.path:
            queryset=queryset.filter(is_active=True).order_by('id')
        return queryset,use_distinct 

    def has_change_permission(self, request, obj=None):
        return False
    def has_delete_permission(self, request, obj=None):
        return False  # 全局禁用删除按钮
    
@admin.register(Company)  
class CompanyAdmin(GlobalAdmin):   
    exclude = ('id','createtime','updatetime','is_active')
    def has_change_permission(self, request, obj=None):
        return False
    def has_delete_permission(self, request, obj=None):
        return False  # 全局禁用删除按钮
    

@admin.register(CompetitionRelation)  
class CompetitionRelationAdmin(GlobalAdmin):   
    search_fields=['competitionrelation']
    exclude = ('id','createtime','updatetime','is_active')


@admin.register(MindrayInstrumentCategory)
class MindrayInstrumentCategoryAdmin(GlobalAdmin):
    list_display = ['name', 'order', 'is_active', 'createtime']
    list_filter = ['is_active']
    ordering = ['order']
    
    def get_queryset(self, request):
        return super().get_queryset(request).filter(is_active=True)

 
class MindrayInstrumentSurveyForm(forms.ModelForm):
    class Meta:
        model = MindrayInstrumentSurvey
        fields = '__all__'
        # widgets = {
        #     'expiry_date': YearMonthDateWidget(attrs={
        #         'placeholder': '选择到期时间'
        #     }),
        # }
    
    # def clean_expiry_date(self):
    #     """确保日期字段正确处理"""
    #     expiry_date = self.cleaned_data.get('expiry_date')
    #     return expiry_date
    
    def clean_quantity(self):
        """验证台数不能小于1"""
        quantity = self.cleaned_data.get('quantity')
        if quantity is not None and quantity < 0:
            raise forms.ValidationError('台数不能小于0')
        return quantity

 
class BloodCellProjectInlineForInstrument(admin.StackedInline):  # 改为StackedInline
    model = MindrayBloodCellProject
    extra = 0
    verbose_name = "血球项目详情"
    verbose_name_plural = "血球项目详情"
    can_delete = True
    
 
    fields = [
        ('project_type', 'sample_volume'), 
        ('competitionrelation', 'dealer_name')
    ]
    autocomplete_fields = ['competitionrelation']
    
    def get_queryset(self, request):
        return super().get_queryset(request).filter(is_active=True)
    
    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "competitionrelation":
            kwargs["queryset"] = CompetitionRelation.objects.filter(is_active=True)
        return super().formfield_for_foreignkey(db_field, request, **kwargs)


class MindrayBloodCellProjectInline(nested_admin.NestedTabularInline):
    model = MindrayBloodCellProject
    extra = 1
    verbose_name = "血球项目"
    verbose_name_plural = "血球项目"
    autocomplete_fields = ['competitionrelation']

    fields = [
        'project_type', 'sample_volume', 
        'competitionrelation', 'dealer_name'
    ]
    
    def get_queryset(self, request):
        return super().get_queryset(request).filter(is_active=True)
    
    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "competitionrelation":
            kwargs["queryset"] = CompetitionRelation.objects.filter(is_active=True)
        return super().formfield_for_foreignkey(db_field, request, **kwargs)
    
 
class BaseInstrumentInline(nested_admin.NestedStackedInline):
    model = MindrayInstrumentSurvey
    form = MindrayInstrumentSurveyForm
    extra = 0
    verbose_name = "仪器调研"
    verbose_name_plural = "仪器调研"
    
    autocomplete_fields = ['brand', 'competitionrelation']
    exclude = ['category']
    
    def get_extra(self, request, obj=None, **kwargs):
        if obj is None:
            return 1
        return 0
    
    def get_max_num(self, request, obj=None, **kwargs):
        return None
    
    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "brand":
            kwargs["queryset"] = Brand.objects.filter(is_active=True)
        elif db_field.name == "competitionrelation":
            kwargs["queryset"] = CompetitionRelation.objects.filter(is_active=True)
        return super().formfield_for_foreignkey(db_field, request, **kwargs)
    
    # 新增这个辅助方法
    def _sync_last_modified_by(self, instance, user):
        """辅助方法：为仪器实例设置修改人（仅当为空时）"""
        if hasattr(instance, 'last_modified_by') and not instance.last_modified_by:
            instance.last_modified_by = user
            return True
        return False
    
class BloodCellInstrumentInline(BaseInstrumentInline):
    verbose_name = "血球仪器调研"
    verbose_name_plural = "血球仪器调研"
    inlines = [MindrayBloodCellProjectInline]
    
    fieldsets = (
        ('基本信息', {
            'fields': (
                ('is_our_instrument', 'our_sales_channel'),
                ('brand', 'model'),
                ('quantity', 'installation_year'),
                'installation_location',
            ),
            'classes': ('primary-fieldset',)
        }),
    )
    
    def get_queryset(self, request):
        return super().get_queryset(request).filter(
            is_active=True,
            category__name="血球"
        )
     
 
    def get_formset(self, request, obj=None, **kwargs):
            formset_class = super().get_formset(request, obj, **kwargs)        
            base_inline = self
    
            class BloodCellFormSet(formset_class):
                def save(self, commit=True):
                    instances = super().save(commit=False)
                    
                    # 检查是否有实际改动
                    has_changes = bool(instances) or bool(self.deleted_objects)
                    
                    try:
                        category = MindrayInstrumentCategory.objects.get(name="血球", is_active=True)
                        for instance in instances:
                            if not instance.category_id:
                                instance.category = category
                            
                            base_inline._sync_last_modified_by(instance, request.user)
                            
                            if commit:
                                instance.save()
                                instance.calculate_all_blood_summaries()
                                instance.save(update_fields=[
                                    'sample_volume', 'blood_project_types', 
                                    'blood_project_details',
                                    'blood_competition_relations', 'blood_dealer_names'
                                    # 注意：这里不包含updatetime，让模型的auto_now处理
                                ])
                    except MindrayInstrumentCategory.DoesNotExist:
                        pass
                    
                    if commit:
                        self.save_m2m()
                        # 只有在有变化时才触发医院统计重新计算
                        if has_changes and instances:
                            hospital_survey = instances[0].hospital_survey
                            hospital_survey.calculate_all_statistics()
                            
                            # 手动更新医院调研的updatetime
                            from django.utils import timezone
                            hospital_survey.updatetime = timezone.now()
                            
                            hospital_survey.save(update_fields=[
                                'crp_total_volume', 'saa_total_volume', 
                                'esr_total_volume', 'routine_total_volume',
                                'glycation_total_volume', 'urine_total_volume',
                                'blood_cell_total_count', 'glycation_total_count', 'urine_total_count',
                                'blood_cell_summary', 'glycation_summary', 'urine_summary',
                                'updatetime'  # 只有在有变化时更新
                            ])
                    
                    return instances
            
            return BloodCellFormSet     


class GlycationInstrumentInline(BaseInstrumentInline):
    verbose_name = "糖化仪器调研"
    verbose_name_plural = "糖化仪器调研"
    
    fieldsets = (
        ('仪器信息', {
            'fields': (
                ('is_our_instrument', 'our_sales_channel'),
                ('brand', 'model'),
                ('quantity', 'installation_year'),
                ('sample_volume', 'installation_location'),
                ('competitionrelation', 'dealer_name'),
            ),
            'classes': ('primary-fieldset',)
        }),
    )
    
    def get_queryset(self, request):
        return super().get_queryset(request).filter(
            is_active=True,
            category__name="糖化"
        )
    
  
    def get_formset(self, request, obj=None, **kwargs):
        formset_class = super().get_formset(request, obj, **kwargs)
        base_inline = self  # 保存引用
        
        class GlycationFormSet(formset_class):
            def save(self, commit=True):
                instances = super().save(commit=False)
                
                # 检查是否有实际改动
                has_changes = bool(instances) or bool(self.deleted_objects)
                
                if not has_changes:
                    return instances  # 没有改动，直接返回
                
                try:
                    category = MindrayInstrumentCategory.objects.get(name="糖化", is_active=True)
                    for instance in instances:
                        if not instance.category_id:
                            instance.category = category
                        
                        # 同步修改人
                        base_inline._sync_last_modified_by(instance, request.user)
                        
                        if commit:
                            instance.save()
                except MindrayInstrumentCategory.DoesNotExist:
                    pass
                
                if commit:
                    self.save_m2m()
                    # 只有在有变化时才触发医院统计重新计算
                    if instances:
                        hospital_survey = instances[0].hospital_survey
                        hospital_survey.calculate_all_statistics()
                        
                        # 手动更新医院调研的updatetime
                        from django.utils import timezone
                        hospital_survey.updatetime = timezone.now()
                        
                        hospital_survey.save(update_fields=[
                            'crp_total_volume', 'saa_total_volume', 
                            'esr_total_volume', 'routine_total_volume',
                            'glycation_total_volume', 'urine_total_volume',
                            'blood_cell_total_count', 'glycation_total_count', 'urine_total_count',
                            'blood_cell_summary', 'glycation_summary', 'urine_summary',
                            'updatetime'  # 只有在有变化时更新
                        ])
                
                return instances
        
        return GlycationFormSet


class UrineInstrumentInline(BaseInstrumentInline):
    verbose_name = "尿液仪器调研"
    verbose_name_plural = "尿液仪器调研"
    
    fieldsets = (
        ('仪器信息', {
            'fields': (
                ('is_our_instrument', 'our_sales_channel'),
                ('brand', 'model'),
                ('quantity', 'installation_year'),
                ('sample_volume', 'installation_location'),
                ('competitionrelation', 'dealer_name'),
            ),
            'classes': ('primary-fieldset',)
        }),
    )
    
    def get_queryset(self, request):
        return super().get_queryset(request).filter(
            is_active=True,
            category__name="尿液"
        )

    def get_formset(self, request, obj=None, **kwargs):
        formset_class = super().get_formset(request, obj, **kwargs)
        base_inline = self  # 保存引用
        
        class UrineFormSet(formset_class):
            def save(self, commit=True):
                instances = super().save(commit=False)
                
                # 检查是否有实际改动
                has_changes = bool(instances) or bool(self.deleted_objects)
                
                if not has_changes:
                    return instances  # 没有改动，直接返回
                
                try:
                    category = MindrayInstrumentCategory.objects.get(name="尿液", is_active=True)
                    for instance in instances:
                        if not instance.category_id:
                            instance.category = category
                        
                        # 同步修改人
                        base_inline._sync_last_modified_by(instance, request.user)
                        
                        if commit:
                            instance.save()
                except MindrayInstrumentCategory.DoesNotExist:
                    pass
                
                if commit:
                    self.save_m2m()
                    # 只有在有变化时才触发医院统计重新计算
                    if instances:
                        hospital_survey = instances[0].hospital_survey
                        hospital_survey.calculate_all_statistics()
                        
                        # 手动更新医院调研的updatetime
                        from django.utils import timezone
                        hospital_survey.updatetime = timezone.now()
                        
                        hospital_survey.save(update_fields=[
                            'crp_total_volume', 'saa_total_volume', 
                            'esr_total_volume', 'routine_total_volume',
                            'glycation_total_volume', 'urine_total_volume',
                            'blood_cell_total_count', 'glycation_total_count', 'urine_total_count',
                            'blood_cell_summary', 'glycation_summary', 'urine_summary',
                            'updatetime'  # 只有在有变化时更新
                        ])
                
                return instances
        
        return UrineFormSet


class SalesOpportunityInline(nested_admin.NestedTabularInline):
    model = SalesOpportunity
    extra = 0
    max_num = 0  # 不允许添加新记录
    can_delete = False  # 不允许删除
    readonly_fields = [
         'opportunity_model', 'opportunity_project', 
        'sample_volume', 'landing_time', 'opportunity_status', 'createtime'
    ]
    fields = readonly_fields
    verbose_name = "相关商机"
    verbose_name_plural = "相关商机列表"
    
    # 添加 nested_admin 需要的属性
    sortable_options = {}
    
    def has_add_permission(self, request, obj=None):
        return False  # 禁止添加
    
    def has_delete_permission(self, request, obj=None):
        return False  # 禁止删除


@admin.register(MindrayHospitalSurvey)
class MindrayHospitalSurveyAdmin(nested_admin.NestedModelAdmin, GlobalAdmin):
    ordering = ['id']
    autocomplete_fields = ['hospital', 'qitian_manager', 'director_contact', 
                          'leader_contact', 'operator_contact']
    
    list_display = [
        'hospital', 
        'get_hospital_district',
        'get_hospital_class',
        'qitian_manager', 
        'mindray_manager',
        'director_name',
        'director_familiarity_colored',
        'leader_name',
        'leader_familiarity_colored', 
        'sales_mode',
        'distribution_channel',
        'get_routine_volume',       # 血常规标本量
        'get_crp_volume',           # CRP标本量
        'get_saa_volume',           # SAA标本量
        'get_esr_volume',           # 血沉标本量

        'get_glycation_volume',     # 新增：糖化标本量
        'get_urine_volume',         # 新增：尿液标本量

        'get_blood_cell_count',     # 血球仪器台数
        'get_glycation_count',      # 糖化仪器台数
        'get_urine_count',          # 尿液仪器台数
        'get_blood_cell_summary',   # 血球仪器汇总
        'get_glycation_summary',    # 新增：糖化仪器汇总
        'get_urine_summary',        # 新增：尿液仪器汇总
        'display_sales_opportunities_summary',  # 新增：显示商机汇总
        'created_by',
        'updatetime'
    ]
    
    list_filter = [
        'hospital__district', 
        'hospital__hospitalclass',
        SalesmanFilter,
        'sales_mode',
        'director_familiarity',
        'leader_familiarity',
    ]
    
    search_fields = [
        'hospital__hospitalname', 
        'mindray_manager', 
        'director_name',
        'leader_name',
        'operator_name'
    ]
    
    
    fieldsets = (
        ('医院基本信息', {
            'fields': ('hospital', 'qitian_manager', 'mindray_manager'),
            'classes': ('primary-fieldset',)
        }),
        ('主任信息', {
            'fields': ('director_name', 'director_contact', 'director_familiarity'),
            'classes': ('secondary-fieldset',)
        }),
        ('组长信息', {
            'fields': ('leader_name', 'leader_contact', 'leader_familiarity'),
            'classes': ('secondary-fieldset',)
        }),
        ('操作老师信息', {
            'fields': ('operator_name', 'operator_contact'),
            'classes': ('secondary-fieldset',)
        }),
        ('销售相关信息', {
            'fields': ('sales_mode', 'distribution_channel'),
            'classes': ('tertiary-fieldset',)
        }),
    
    )
    
    readonly_fields = [
        'crp_total_volume', 'saa_total_volume', 'esr_total_volume', 'routine_total_volume',
        'glycation_total_volume', 'urine_total_volume',  # 新增：只读字段
        'blood_cell_total_count', 'glycation_total_count', 'urine_total_count',
        'blood_cell_summary', 'glycation_summary', 'urine_summary',# 新增后两个字段
        'sales_opportunities_summary'
    ]

    inlines = [BloodCellInstrumentInline, GlycationInstrumentInline, UrineInstrumentInline,SalesOpportunityInline]
    
    # 原有显示方法保持不变...
    def get_crp_volume(self, obj):
        return f"{obj.crp_total_volume}"
    get_crp_volume.short_description = 'CRP标本量'
    get_crp_volume.admin_order_field = 'crp_total_volume'
    
    def get_saa_volume(self, obj):
        return f"{obj.saa_total_volume}"
    get_saa_volume.short_description = 'SAA标本量'
    get_saa_volume.admin_order_field = 'saa_total_volume'
    
    def get_esr_volume(self, obj):
        return f"{obj.esr_total_volume}"
    get_esr_volume.short_description = '血沉标本量'
    get_esr_volume.admin_order_field = 'esr_total_volume'
    
    def get_routine_volume(self, obj):
        return f"{obj.routine_total_volume}"
    get_routine_volume.short_description = '血常规标本量'
    get_routine_volume.admin_order_field = 'routine_total_volume'
    
     
    def get_blood_cell_count(self, obj):
        return f"{obj.blood_cell_total_count}"
    get_blood_cell_count.short_description = '血球台数'
    get_blood_cell_count.admin_order_field = 'blood_cell_total_count'
    
    def get_glycation_count(self, obj):
        return f"{obj.glycation_total_count}"
    get_glycation_count.short_description = '糖化台数'
    get_glycation_count.admin_order_field = 'glycation_total_count'
    
    def get_urine_count(self, obj):
        return f"{obj.urine_total_count}"
    get_urine_count.short_description = '尿液台数'
    get_urine_count.admin_order_field = 'urine_total_count'
    
    # def get_blood_cell_summary(self, obj):
    #     """显示血球仪器汇总，超长时截断"""
    #     summary = obj.blood_cell_summary or "无"
    #     # if len(summary) > 100:  # 如果太长，截断显示
    #     #     return f"{summary[:100]}..."
    #     return summary
    # get_blood_cell_summary.short_description = '血球品牌-型号-台数-装机年份-标本量'
    # get_blood_cell_summary.admin_order_field = 'blood_cell_summary'
    
 
    # def get_glycation_summary(self, obj):
    #     """显示糖化仪器汇总，超长时截断"""
    #     summary = obj.glycation_summary or "无"
    #     # if len(summary) > 100:  # 如果太长，截断显示
    #     #     return f"{summary[:100]}..."
    #     return summary
    # get_glycation_summary.short_description = '糖化品牌-型号-台数-装机年份-标本量'
    # get_glycation_summary.admin_order_field = 'glycation_summary'
    
    # def get_urine_summary(self, obj):
    #     """显示尿液仪器汇总，超长时截断"""
    #     summary = obj.urine_summary or "无"
    #     # if len(summary) > 100:  # 如果太长，截断显示
    #     #     return f"{summary[:100]}..."
    #     return summary
    # get_urine_summary.short_description = '尿液品牌-型号-台数-装机年份-标本量'
    # get_urine_summary.admin_order_field = 'urine_summary'

    def get_blood_cell_summary(self, obj):
        """显示血球仪器汇总，支持换行显示"""
        summary = obj.blood_cell_summary or "无"
        # 将换行符转换为HTML换行标签
        if summary and summary != "无":
            summary_html = summary.replace('\n', '<br>')
            return format_html(summary_html)
        return summary

    get_blood_cell_summary.short_description = '血球品牌-型号-台数-装机年份-标本量'
    get_blood_cell_summary.admin_order_field = 'blood_cell_summary'

    def get_glycation_summary(self, obj):
        """显示糖化仪器汇总，支持换行显示"""
        summary = obj.glycation_summary or "无"
        # 将换行符转换为HTML换行标签
        if summary and summary != "无":
            summary_html = summary.replace('\n', '<br>')
            return format_html(summary_html)
        return summary

    get_glycation_summary.short_description = '糖化品牌-型号-台数-装机年份-标本量'
    get_glycation_summary.admin_order_field = 'glycation_summary'

    def get_urine_summary(self, obj):
        """显示尿液仪器汇总，支持换行显示"""
        summary = obj.urine_summary or "无"
        # 将换行符转换为HTML换行标签
        if summary and summary != "无":
            summary_html = summary.replace('\n', '<br>')
            return format_html(summary_html)
        return summary

    get_urine_summary.short_description = '尿液品牌-型号-台数-装机年份-标本量'
    get_urine_summary.admin_order_field = 'urine_summary'


    # def display_sales_opportunities_summary(self, obj):
    #     """在列表页显示商机汇总，支持换行显示"""
    #     if obj.sales_opportunities_summary and obj.sales_opportunities_summary != "无":
    #         # 将/分隔的内容用<br>标签连接，实现换行显示
    #         summaries = obj.sales_opportunities_summary.split('/')
    #         html_content = '<br>'.join(summaries)
    #         return format_html(html_content)
    #     return "无商机"
    
    # display_sales_opportunities_summary.short_description = '商机项目-型号-标本量-落地时间'
    # display_sales_opportunities_summary.admin_order_field = 'sales_opportunities_summary'

    def display_sales_opportunities_summary(self, obj):
        """显示尿液仪器汇总，支持换行显示"""
        summary = obj.sales_opportunities_summary or "无"
        # 将换行符转换为HTML换行标签
        if summary and summary != "无":
            summary_html = summary.replace('\n', '<br>')
            return format_html(summary_html)
        return summary

    display_sales_opportunities_summary.short_description = '商机项目-型号-标本量-落地时间'
    display_sales_opportunities_summary.admin_order_field = 'sales_opportunities_summary'

#====================================================


    def get_hospital_district(self, obj):
        return obj.hospital.district
    get_hospital_district.short_description = '区域'
    get_hospital_district.admin_order_field = 'hospital__district'
    
    def get_hospital_class(self, obj):
        return obj.hospital.hospitalclass  
    get_hospital_class.short_description = '级别'
    get_hospital_class.admin_order_field = 'hospital__hospitalclass'
    
       
    def director_familiarity_colored(self, obj):
        color_map = {
            'red': "#f40909",
            'yellow': "#ffcc00",
            'green': "#077e07",
            'blue':"#3056d2",
        }
        color = color_map.get(obj.director_familiarity, '#000000')
        text_map = {
            'red': '不认识',
            'yellow': '有商机在跟进',
            'green': '有明确代理商',
            'blue':'成单'
        }
        text = text_map.get(obj.director_familiarity, obj.director_familiarity)
        return format_html(
            '<span style="color: {}; font-weight: bold;">{}</span>',
            color, text
        )
    director_familiarity_colored.short_description = '主任客情'
    director_familiarity_colored.admin_order_field = 'director_familiarity'
    
    def leader_familiarity_colored(self, obj):
        color_map = {
            'red': "#f40909",
            'yellow': "#ffcc00",
            'green': "#077e07",
            'blue':"#3056d2",
        }
        color = color_map.get(obj.leader_familiarity, '#000000')
        text_map = {
            'red': '不认识',
            'yellow': '有商机在跟进',
            'green': '有明确代理商',
            'blue':'成单'
        }
        text = text_map.get(obj.leader_familiarity, obj.leader_familiarity)
        return format_html(
            '<span style="color: {}; font-weight: bold;">{}</span>',
            color, text
        )
    leader_familiarity_colored.short_description = '组长客情'
    leader_familiarity_colored.admin_order_field = 'leader_familiarity'
    

    # 新增显示方法
    def get_glycation_volume(self, obj):
        return f"{obj.glycation_total_volume}"
    get_glycation_volume.short_description = '糖化标本量'
    get_glycation_volume.admin_order_field = 'glycation_total_volume'
    
    def get_urine_volume(self, obj):
        return f"{obj.urine_total_volume}"
    get_urine_volume.short_description = '尿液标本量'
    get_urine_volume.admin_order_field = 'urine_total_volume'


    def get_queryset(self, request):
        return super().get_queryset(request).filter(is_active=True)
    
    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "hospital":
            kwargs["queryset"] = Hospital.objects.filter(is_active=True)
        elif db_field.name in ["qitian_manager", "director_contact", "leader_contact", "operator_contact", "created_by"]:
            kwargs["queryset"] = UserInfoMindray.objects.filter(is_active=True)
        return super().formfield_for_foreignkey(db_field, request, **kwargs)
     
    # 新增这个方法
    def _sync_last_modified_by_to_instruments(self, hospital_survey, user):
        """将修改人同步到相关仪器调研记录（仅当仪器记录的last_modified_by为空时）"""
        try:
            updated_count = MindrayInstrumentSurvey.objects.filter(
                hospital_survey=hospital_survey,
                is_active=True,
                last_modified_by__isnull=True
            ).update(last_modified_by=user)
            
            if updated_count > 0:
                print(f"已同步修改人到 {updated_count} 条仪器调研记录")
        except Exception as e:
            print(f"同步修改人到仪器调研时出错: {e}")
    

    def save_model(self, request, obj, form, change):
        """只有在真正有改动时才更新 updatetime"""
        if not change:
            obj.created_by = request.user
        else:
            # 只在修改时更新created_by作为最后修改人
            obj.created_by = request.user
        
        # 检查是否有实际字段改动
        has_changes = False
        if change and form.changed_data:
            has_changes = True
        
        # 如果没有改动，使用update_fields避免触发auto_now
        if change and not has_changes:
            # 没有改动时，不保存（或者使用refresh_from_db）
            obj.refresh_from_db()
            return
        
        super().save_model(request, obj, form, change)
        
        # 只有在有改动时才同步修改人到仪器记录
        if has_changes or not change:
            self._sync_last_modified_by_to_instruments(obj, request.user)
    
    def save_formset(self, request, form, formset, change):
        """改进的save_formset方法，只有实际变化时才更新统计和时间"""
        # 检查formset是否有实际改动
        has_formset_changes = False
        
        instances = formset.save(commit=False)
        
        # 检查新增、修改的实例
        for instance in instances:
            instance.hospital_survey = form.instance
            instance.full_clean()
            instance.save()
            has_formset_changes = True
        
        # 检查删除的实例
        for obj in formset.deleted_objects:
            obj.is_active = False
            obj.save()
            has_formset_changes = True
        
        formset.save_m2m()
        
        # 只有在formset有变化时才重新计算统计并更新时间
        if has_formset_changes:
            # 保存完成后重新计算所有统计数据
            form.instance.calculate_all_statistics()
            
            # 手动更新updatetime（因为我们需要显式控制）
            from django.utils import timezone
            form.instance.updatetime = timezone.now()
            
            form.instance.save(update_fields=[
                'crp_total_volume', 'saa_total_volume', 
                'esr_total_volume', 'routine_total_volume',
                'glycation_total_volume', 'urine_total_volume',
                'blood_cell_total_count', 'glycation_total_count', 'urine_total_count',
                'blood_cell_summary', 'glycation_summary', 'urine_summary','sales_opportunities_summary',
                'updatetime'  # 只有在有变化时才更新
            ])
            
            # 只在有变化时才同步修改人
            self._sync_last_modified_by_to_instruments(form.instance, request.user)

    # 更新actions
    actions = ['download_excel','download_comprehensive_workbook','refresh_all_calculated_fields']
    

    def download_excel(self, request, queryset):
        """导出选中的医院调研数据到Excel"""
        # 如果没有选中任何项，则导出当前页面的所有数据
        if not queryset.exists():
            queryset = self.get_queryset(request)
        
        # 创建工作簿和工作表
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "医院调研数据"
        
        # 定义表头 - 修改这部分
        headers = [
            '医院名称', '区域', '级别', '其田负责人', '迈瑞负责人',
            '主任姓名','主任对接人','主任客情',
            '组长姓名','组长对接人','组长客情', 
            '销售模式', '分销渠道',
            '血常规标本量','CRP标本量', 'SAA标本量', '血沉标本量',
            '糖化标本量', '尿液标本量',  # 新增这两个字段
            '血球台数', '糖化台数', '尿液台数',
            '血球品牌-型号-台数-装机年份-标本量',
            '糖化品牌-型号-台数-装机年份-标本量', 
            '尿液品牌-型号-台数-装机年份-标本量',
            '商机型号-项目-标本量-落地时间',  # 新增：商机汇总列
            '修改人', '更新时间'
        ]
        
        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 写入表头
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 客情显示映射
        familiarity_map = {
            'red': '不认识',
            'yellow': '有商机在跟进',
            'green': '有明确代理商',
            'blue': '成单'
        }
        
     
        familiarity_color_map = {
            'red':    'FFFFE6E6',
            'yellow': 'FFFFFFCC',
            'green':  'FFE6FFE6',
            'blue':   'FFE6F2FF'
        }
        familiarity_font_color_map = {
            'red':    'FFF40909',
            'yellow': 'FFCC9900',
            'green':  'FF077E07',
            'blue':   'FF3056D2'
        }



        # 写入数据
        for row_num, obj in enumerate(queryset, 2):
            # 确保在导出前计算最新的统计数据
            obj.calculate_all_statistics()
            
            # 处理更新时间 - 关键修改在这里
            update_time_str = ''
            if obj.updatetime:
                # 将UTC时间转换为本地时间
                local_time = timezone.localtime(obj.updatetime)
                update_time_str = local_time.strftime('%Y-%m-%d %H:%M:%S')

            row_data = [
                obj.hospital.hospitalname if obj.hospital else '',
                obj.hospital.district if obj.hospital else '',
                obj.hospital.hospitalclass if obj.hospital else '',
                obj.qitian_manager.chinesename if obj.qitian_manager else '',
                obj.mindray_manager or '',                
                obj.director_name or '',
                obj.director_contact.chinesename if obj.director_contact else '',
                familiarity_map.get(obj.director_familiarity, obj.director_familiarity or ''),
                obj.leader_name or '',
                obj.leader_contact.chinesename if obj.leader_contact else '',
                familiarity_map.get(obj.leader_familiarity, obj.leader_familiarity or ''),
                dict(obj.SALES_MODE_CHOICES).get(obj.sales_mode, obj.sales_mode or ''),
                obj.distribution_channel or '',
                obj.routine_total_volume,   
                obj.crp_total_volume,
                obj.saa_total_volume,
                obj.esr_total_volume,      
                obj.glycation_total_volume,  # 新增：糖化标本量
                obj.urine_total_volume,      # 新增：尿液标本量          
                obj.blood_cell_total_count,
                obj.glycation_total_count,
                obj.urine_total_count,
                obj.blood_cell_summary or '',
                obj.glycation_summary or '',
                obj.urine_summary or '',
                obj.sales_opportunities_summary or '' ,
                obj.created_by.chinesename if obj.created_by else '',
                update_time_str
            ]
            
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        
               # 再精准获取要上色的单元格，仅第 8、11 列
            def paint_cell(r, c, key):
                bg = familiarity_color_map.get(key)
                fg = familiarity_font_color_map.get(key)
                if not (bg and fg):
                    return
                target = ws.cell(row=r, column=c)
                # 使用 8 位 ARGB 颜色更稳（前缀 FF 表示不透明）
                if len(bg) == 6: bg = 'FF' + bg
                if len(fg) == 6: fg = 'FF' + fg
                target.fill = PatternFill(start_color=bg.upper(), end_color=bg.upper(), fill_type="solid")
                target.font = Font(color=fg.upper(), bold=True)
                target.alignment = Alignment(horizontal="center", vertical="center")

            if obj.director_familiarity:
                paint_cell(row_num, 8, obj.director_familiarity)

            if obj.leader_familiarity:
                paint_cell(row_num, 11, obj.leader_familiarity)


        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            # 设置最小宽度10，最大宽度50
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 冻结首行
        ws.freeze_panes = 'A2'
        
        # 创建HTTP响应
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # 设置文件名（包含当前日期时间）
        current_time = timezone.now().strftime('%Y%m%d_%H%M%S')
        filename = f'医院调研数据_{current_time}.xlsx'

        # 方法1：使用URL编码（推荐）
        encoded_filename = quote(filename.encode('utf-8'))
        response['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{encoded_filename}'
         
        # 保存工作簿到响应
        wb.save(response)
        
        return response
    
    download_excel.short_description = "导出列表"
    download_excel.type = 'warning'
    download_excel.style = 'color:white;'


    def download_comprehensive_workbook(self, request, queryset):
        """导出综合工作簿（包含5个工作表）"""
        # 如果没有选中任何项，则导出当前页面的所有数据
        if not queryset.exists():
            queryset = self.get_queryset(request)
            # 应用搜索和过滤条件
            cl = self.get_changelist_instance(request)
            queryset = cl.get_queryset(request)
        
        # 创建工作簿
        wb = openpyxl.Workbook()
        # 删除默认的工作表
        wb.remove(wb.active)
        
        # 通用样式定义
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 客情颜色映射
        familiarity_map = {
            'red': '不认识',
            'yellow': '有商机在跟进',
            'green': '有明确代理商',
            'blue': '成单'
        }
        familiarity_color_map = {
            'red': 'FFFFE6E6',
            'yellow': 'FFFFFFCC',
            'green': 'FFE6FFE6',
            'blue': 'FFE6F2FF'
        }
        familiarity_font_color_map = {
            'red': 'FFF40909',
            'yellow': 'FFCC9900',
            'green': 'FF077E07',
            'blue': 'FF3056D2'
        }
        
        # ==================== Sheet 1: 医院调研数据 ====================
        ws1 = wb.create_sheet(title="调研汇总")
        
        # Sheet1 表头
        headers1 = [
            '医院名称', '区域', '级别', '其田负责人', '迈瑞负责人',
            '主任姓名','主任对接人','主任客情',
            '组长姓名','组长对接人','组长客情', 
            '销售模式', '分销渠道',
            '血常规标本量','CRP标本量', 'SAA标本量', '血沉标本量',
            '糖化标本量', '尿液标本量',
            '血球台数', '糖化台数', '尿液台数',
            '血球品牌-型号-台数-装机年份-标本量',
            '糖化品牌-型号-台数-装机年份-标本量', 
            '尿液品牌-型号-台数-装机年份-标本量',
            '商机型号-项目-标本量-落地时间',
            '修改人', '更新时间'
        ]
        
        # 写入Sheet1表头
        for col_num, header in enumerate(headers1, 1):
            cell = ws1.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 写入Sheet1数据
        for row_num, obj in enumerate(queryset, 2):
            obj.calculate_all_statistics()
            
            update_time_str = ''
            if obj.updatetime:
                local_time = timezone.localtime(obj.updatetime)
                update_time_str = local_time.strftime('%Y-%m-%d %H:%M:%S')

            row_data = [
                obj.hospital.hospitalname if obj.hospital else '',
                obj.hospital.district if obj.hospital else '',
                obj.hospital.hospitalclass if obj.hospital else '',
                obj.qitian_manager.chinesename if obj.qitian_manager else '',
                obj.mindray_manager or '',                
                obj.director_name or '',
                obj.director_contact.chinesename if obj.director_contact else '',
                familiarity_map.get(obj.director_familiarity, obj.director_familiarity or ''),
                obj.leader_name or '',
                obj.leader_contact.chinesename if obj.leader_contact else '',
                familiarity_map.get(obj.leader_familiarity, obj.leader_familiarity or ''),
                dict(obj.SALES_MODE_CHOICES).get(obj.sales_mode, obj.sales_mode or ''),
                obj.distribution_channel or '',
                obj.routine_total_volume,   
                obj.crp_total_volume,
                obj.saa_total_volume,
                obj.esr_total_volume,      
                obj.glycation_total_volume,
                obj.urine_total_volume,            
                obj.blood_cell_total_count,
                obj.glycation_total_count,
                obj.urine_total_count,
                obj.blood_cell_summary or '',
                obj.glycation_summary or '',
                obj.urine_summary or '',
                obj.sales_opportunities_summary or '',
                obj.created_by.chinesename if obj.created_by else '',
                update_time_str
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws1.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 设置客情颜色
            def paint_cell(r, c, key):
                bg = familiarity_color_map.get(key)
                fg = familiarity_font_color_map.get(key)
                if not (bg and fg):
                    return
                target = ws1.cell(row=r, column=c)
                if len(bg) == 6: bg = 'FF' + bg
                if len(fg) == 6: fg = 'FF' + fg
                target.fill = PatternFill(start_color=bg.upper(), end_color=bg.upper(), fill_type="solid")
                target.font = Font(color=fg.upper(), bold=True)
                target.alignment = Alignment(horizontal="center", vertical="center")

            if obj.director_familiarity:
                paint_cell(row_num, 8, obj.director_familiarity)
            if obj.leader_familiarity:
                paint_cell(row_num, 11, obj.leader_familiarity)
        
        # 自动调整Sheet1列宽
        self._auto_adjust_column_width(ws1)
        ws1.freeze_panes = 'A2'
        
        # ==================== Sheet 2: 仪器调研简单数据 ====================
        ws2 = wb.create_sheet(title="仪器明细列表")
        
        # 获取所有仪器调研数据
        instrument_queryset = MindrayInstrumentSurvey.objects.filter(
            is_active=True,
            hospital_survey__in=queryset
        ).select_related(
            'hospital_survey__qitian_manager',
            'hospital_survey__hospital',
            'category',
            'brand',
            'last_modified_by'
        ).order_by('hospital_survey__id', 'category__id', 'brand__id','createtime')
        
        # Sheet2 表头
        headers2 = [
            '医院调研',
            '区域', 
            '级别',
            '医院负责人',
            '仪器分类',
            '是否我司仪器',
            '我司业务销售渠道',
            '品牌',
            '型号',
            '台数',
            '装机年份',
            '仪器安装地',
            '标本量总和',
            '血球项目-标本量-竞品关系点-经销商',
            '竞品关系点',
            '经销商名称',
            '最后修改人',
            '更新时间'
        ]
        
        # 写入Sheet2表头
        for col_num, header in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 写入Sheet2数据
        for row_num, obj in enumerate(instrument_queryset, 2):
            if obj.is_blood_category:
                obj.calculate_all_blood_summaries()
            
            # 计算标本量总和
            total_sample_volume = 0
            if obj.category and obj.category.name == '血球':
                from django.db.models import Sum
                total_sample_volume = MindrayBloodCellProject.objects.filter(
                    instrument_survey=obj,
                    is_active=True
                ).aggregate(total=Sum('sample_volume'))['total'] or 0
            else:
                total_sample_volume = obj.sample_volume or 0
            
            # 统一竞品关系点显示
            competition_relation_display = ''
            if obj.category and obj.category.name == '血球':
                competition_relation_display = obj.blood_competition_relations or '-'
            else:
                competition_relation_display = obj.competitionrelation.competitionrelation if obj.competitionrelation else '-'
            
            # 统一经销商显示
            dealer_name_display = ''
            if obj.category and obj.category.name == '血球':
                dealer_name_display = obj.blood_dealer_names or '-'
            else:
                dealer_name_display = obj.dealer_name or '-'
            
            row_data = [
                str(obj.hospital_survey) if obj.hospital_survey else '',
                obj.hospital_survey.hospital.district if obj.hospital_survey and obj.hospital_survey.hospital else '-',
                obj.hospital_survey.hospital.hospitalclass if obj.hospital_survey and obj.hospital_survey.hospital else '-',
                obj.hospital_survey.qitian_manager.chinesename if obj.hospital_survey and obj.hospital_survey.qitian_manager else '-',
                str(obj.category) if obj.category else '',
                '是' if obj.is_our_instrument else '否',
                obj.get_our_sales_channel_display() or '',
                str(obj.brand) if obj.brand else '',
                obj.model or '',
                obj.quantity or 0,
                obj.installation_year or '',
                obj.get_installation_location_display() or '',  
                total_sample_volume,
                obj.blood_project_details or '' if obj.category and obj.category.name == '血球' else '',
                competition_relation_display,
                dealer_name_display,
                obj.last_modified_by.chinesename if obj.last_modified_by else '-',
                self._format_datetime(obj.updatetime)
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws2.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        self._auto_adjust_column_width(ws2)
        ws2.freeze_panes = 'A2'
        
        # ==================== Sheet 3: 综合仪器调研数据（带合并） ====================
        ws3 = wb.create_sheet(title="调研汇总(含合并单元格)")
        
        # 获取所有医院调研记录
        hospital_surveys_queryset = queryset
        
        # 按医院分组数据
        data_by_hospital = self._group_all_hospital_data_for_sheet3(hospital_surveys_queryset)
        
        # Sheet3 表头（完整版）
        headers3 = [
            '医院', '区域', '级别', '其田负责人', '迈瑞负责人',
            '主任姓名', '主任对接人', '主任客情',
            '组长姓名', '组长对接人', '组长客情',
            '操作老师姓名', '操作老师对接人',
            '销售模式', '分销渠道',
            '血常规标本量', 'CRP标本量', 'SAA标本量', '血沉标本量',
            '糖化标本量', '尿液标本量',
            '血球台数', '糖化台数', '尿液台数',
            '血球品牌-型号-台数-装机年份-标本量',
            '糖化品牌-型号-台数-装机年份-标本量',
            '尿液品牌-型号-台数-装机年份-标本量',
            '仪器分类',
            '是否我司仪器', '我司销售渠道',
            '品牌', '型号', '台数', '装机年份', '安装地点',
            '标本量总和', '血球项目-标本量-竞品关系点-经销商',
            '竞品关系点', '经销商名称',
            '血球项目类型', '血球项目标本量', '血球项目竞品关系点', '血球项目经销商',
            '最后修改人', '更新时间'
        ]
        
        # 写入Sheet3表头
        for col_num, header in enumerate(headers3, 1):
            cell = ws3.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        current_row = 2
        merge_ranges = []
        
        for hospital_survey, categories_data in data_by_hospital.items():
            hospital_start_row = current_row
            
            if not categories_data:
                # 医院没有仪器调研
                self._write_data_row_for_sheet3(ws3, current_row, hospital_survey, None, None, familiarity_map, thin_border)
                current_row += 1
            else:
                # 有仪器调研的情况
                for category, instruments_data in categories_data.items():
                    category_start_row = current_row
                    
                    for instrument, projects_data in instruments_data.items():
                        instrument_start_row = current_row
                        
                        if projects_data:
                            for project in projects_data:
                                self._write_data_row_for_sheet3(ws3, current_row, hospital_survey, instrument, project, familiarity_map, thin_border)
                                current_row += 1
                        else:
                            self._write_data_row_for_sheet3(ws3, current_row, hospital_survey, instrument, None, familiarity_map, thin_border)
                            current_row += 1
                        
                        # 仪器信息合并范围
                        if len(projects_data) > 1:
                            instrument_end_row = current_row - 1
                            merge_ranges.append({
                                'start_row': instrument_start_row,
                                'end_row': instrument_end_row,
                                'start_col': 29,
                                'end_col': 39,
                                'type': 'instrument'
                            })
                    
                    # 仪器分类合并范围
                    total_rows_in_category = sum(len(projects) if projects else 1 for projects in instruments_data.values())
                    if total_rows_in_category > 1:
                        category_end_row = current_row - 1
                        merge_ranges.append({
                            'start_row': category_start_row,
                            'end_row': category_end_row,
                            'start_col': 28,
                            'end_col': 28,
                            'type': 'category'
                        })
            
            # 医院信息合并范围
            total_rows_in_hospital = 1 if not categories_data else sum(
                sum(len(projects) if projects else 1 for projects in instruments_data.values()) 
                for instruments_data in categories_data.values()
            )
            
            if total_rows_in_hospital > 1:
                hospital_end_row = current_row - 1
                merge_ranges.append({
                    'start_row': hospital_start_row,
                    'end_row': hospital_end_row,
                    'start_col': 1,
                    'end_col': 27,
                    'type': 'hospital'
                })
        
        # 执行合并单元格
        self._merge_cells_for_sheet3(ws3, merge_ranges, familiarity_color_map, familiarity_font_color_map)
        
        # 重新应用客情颜色
        self._reapply_familiarity_colors_for_sheet3(ws3, data_by_hospital, familiarity_map, familiarity_color_map, familiarity_font_color_map)
        
        self._auto_adjust_column_width(ws3)
        ws3.freeze_panes = 'A2'
        
  
        # ==================== Sheet 4: 销售商机数据 ====================
        ws4 = wb.create_sheet(title="商机明细")

        # 获取相关的销售商机数据 - 修改排序方式
        opportunity_queryset = SalesOpportunity.objects.filter(
            is_active=True,
            hospital_survey__in=queryset
        ).select_related(
            'hospital_survey__hospital',
            'hospital_survey__qitian_manager', 
            'salesperson'
        ).order_by('hospital_survey__id','opportunity_project', '-createtime')  # 先按hospital_survey的id排序，再按创建时间倒序

        # Sheet4 表头
        headers4 = [
            '医院名称', '区域', '级别', '销售人员', 
            '商机型号', '商机项目', '商机标本量', '落地时间', '商机情况',
            '创建时间', '更新时间'
        ]

        # 写入Sheet4表头
        for col_num, header in enumerate(headers4, 1):
            cell = ws4.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 写入Sheet4数据
        for row_num, obj in enumerate(opportunity_queryset, 2):
            create_time_str = ''
            update_time_str = ''
            if obj.createtime:
                local_create_time = timezone.localtime(obj.createtime)
                create_time_str = local_create_time.strftime('%Y-%m-%d %H:%M:%S')
            if obj.updatetime:
                local_update_time = timezone.localtime(obj.updatetime)
                update_time_str = local_update_time.strftime('%Y-%m-%d %H:%M:%S')
            
            row_data = [
                obj.hospital_survey.hospital.hospitalname if obj.hospital_survey and obj.hospital_survey.hospital else '',
                obj.hospital_survey.hospital.district if obj.hospital_survey and obj.hospital_survey.hospital else '',
                obj.hospital_survey.hospital.hospitalclass if obj.hospital_survey and obj.hospital_survey.hospital else '',
                obj.hospital_survey.qitian_manager.chinesename if obj.hospital_survey and obj.hospital_survey.qitian_manager else '',
                obj.opportunity_model or '',
                obj.get_opportunity_project_display(),
                obj.sample_volume or 0,
                obj.landing_time or '',
                obj.opportunity_status or '',
                create_time_str,
                update_time_str
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws4.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

        self._auto_adjust_column_width(ws4)
        ws4.freeze_panes = 'A2'

        # ==================== Sheet 5: 商机汇总数据 ====================        
         
        ws5 = wb.create_sheet(title="商机汇总")

        # 按照与SalesOpportunityAdmin相同的排序规则排序
        opportunity_queryset = opportunity_queryset.order_by('hospital_survey__id', 'opportunity_project', '-createtime')

        # 获取所有不同的落地时间，按时间排序
        landing_times = sorted(list(set(
            obj.landing_time for obj in opportunity_queryset 
            if obj.landing_time
        )))

        # 构建数据结构：按医院、商机型号、商机项目分组
        summary_data = {}

        for obj in opportunity_queryset:
            if not (obj.hospital_survey and obj.hospital_survey.hospital):
                continue
                
            hospital = obj.hospital_survey.hospital
            hospital_key = (
                hospital.hospitalname,
                hospital.district or '',
                hospital.hospitalclass or '',
                obj.hospital_survey.qitian_manager.chinesename if obj.hospital_survey.qitian_manager else ''
            )
            
            project_key = (
                obj.opportunity_model or '',
                obj.get_opportunity_project_display() or '',
                obj.opportunity_status or '',  # 添加商机情况
                obj.createtime.strftime('%Y-%m-%d') if obj.createtime else ''
            )
            
            full_key = hospital_key + project_key
            
            if full_key not in summary_data:
                summary_data[full_key] = {
                    'hospital_name': hospital_key[0],
                    'district': hospital_key[1], 
                    'hospital_class': hospital_key[2],
                    'salesperson': hospital_key[3],
                    'opportunity_model': project_key[0],
                    'opportunity_project': project_key[1],
                    'opportunity_status': project_key[2],  # 添加商机情况
                    'create_time': project_key[3],  # 索引位置调整
                    'landing_volumes': {},
                    'total_volume': 0,
                    # 添加排序辅助字段
                    'hospital_survey_id': obj.hospital_survey.id if obj.hospital_survey else 0,
                    'opportunity_project_order': obj.opportunity_project,
                    'createtime': obj.createtime
                }
            
            # 按落地时间汇总标本量
            landing_time = obj.landing_time or '未设置'
            volume = obj.sample_volume or 0
            
            if landing_time not in summary_data[full_key]['landing_volumes']:
                summary_data[full_key]['landing_volumes'][landing_time] = 0
            
            summary_data[full_key]['landing_volumes'][landing_time] += volume
            summary_data[full_key]['total_volume'] += volume

        # 对汇总数据按照相同的排序规则排序
        sorted_items = sorted(summary_data.items(), key=lambda x: (
            x[1]['hospital_survey_id'],
            x[1]['opportunity_project_order'] or '',
            x[1]['createtime'] or timezone.now()
        ))

        # 构建表头 - 添加商机情况列
        base_headers = ['医院', '区域', '级别', '销售人员', '商机型号', '商机项目', '商机情况', '创建时间']
        landing_time_headers = [str(lt) for lt in landing_times]
        if '未设置' not in landing_time_headers:
            landing_time_headers.append('未设置')

        all_headers = base_headers + landing_time_headers + ['总计']

        # 写入表头
        for col_num, header in enumerate(all_headers, 1):
            cell = ws5.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 按医院分组数据，用于合并单元格 - 使用排序后的数据
        hospital_groups = {}
        for key, data in sorted_items:
            hospital_info = (data['hospital_name'], data['district'], 
                            data['hospital_class'], data['salesperson'])
            if hospital_info not in hospital_groups:
                hospital_groups[hospital_info] = []
            hospital_groups[hospital_info].append((key, data))

        # 定义居中对齐样式
        center_alignment = Alignment(horizontal="center", vertical="center")

        # 写入数据
        current_row = 2
        total_by_landing_time = {}  # 用于计算各落地时间的总计
        grand_total = 0  # 总计

        # 按排序后的医院顺序处理
        processed_hospitals = set()
        for key, data in sorted_items:
            hospital_info = (data['hospital_name'], data['district'], 
                            data['hospital_class'], data['salesperson'])
            
            if hospital_info in processed_hospitals:
                continue
            
            processed_hospitals.add(hospital_info)
            projects = hospital_groups[hospital_info]
            hospital_start_row = current_row
            
            # 写入该医院的所有项目数据
            for proj_key, proj_data in projects:
                # 基础信息列 - 添加居中对齐
                cell1 = ws5.cell(row=current_row, column=1, value=proj_data['hospital_name'])
                cell1.border = thin_border
                cell1.alignment = center_alignment
                
                cell2 = ws5.cell(row=current_row, column=2, value=proj_data['district'])
                cell2.border = thin_border
                cell2.alignment = center_alignment
                
                cell3 = ws5.cell(row=current_row, column=3, value=proj_data['hospital_class'])
                cell3.border = thin_border
                cell3.alignment = center_alignment
                
                cell4 = ws5.cell(row=current_row, column=4, value=proj_data['salesperson'])
                cell4.border = thin_border
                cell4.alignment = center_alignment
                
                ws5.cell(row=current_row, column=5, value=proj_data['opportunity_model']).border = thin_border
                ws5.cell(row=current_row, column=6, value=proj_data['opportunity_project']).border = thin_border
                ws5.cell(row=current_row, column=7, value=proj_data['opportunity_status']).border = thin_border  # 添加商机情况
                ws5.cell(row=current_row, column=8, value=proj_data['create_time']).border = thin_border  # 列数后移
                
                # 各落地时间的标本量 - 起始列数调整为9
                col_num = 9
                for landing_time in landing_time_headers:
                    volume = proj_data['landing_volumes'].get(landing_time, 0)
                    cell = ws5.cell(row=current_row, column=col_num, value=volume if volume > 0 else None)
                    cell.border = thin_border
                    
                    # 累加到总计中
                    if landing_time not in total_by_landing_time:
                        total_by_landing_time[landing_time] = 0
                    total_by_landing_time[landing_time] += volume
                    
                    col_num += 1
                
                # 该项目总计
                cell = ws5.cell(row=current_row, column=len(all_headers), value=proj_data['total_volume'])
                cell.border = thin_border
                grand_total += proj_data['total_volume']
                
                current_row += 1
            
            # 合并医院相关的单元格（1-4列）
            if len(projects) > 1:
                hospital_end_row = current_row - 1
                for col in range(1, 5):  # 列1-4：医院、区域、级别、销售人员
                    ws5.merge_cells(
                        start_row=hospital_start_row, 
                        start_column=col,
                        end_row=hospital_end_row, 
                        end_column=col
                    )
                    # 设置合并单元格的对齐方式
                    merged_cell = ws5.cell(row=hospital_start_row, column=col)
                    merged_cell.alignment = center_alignment

        # 添加总计行
        total_row = current_row + 1
        total_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

        cell = ws5.cell(row=total_row, column=1, value="总计")
        cell.font = Font(bold=True)
        cell.fill = total_fill
        cell.border = thin_border
        cell.alignment = center_alignment  # 为总计单元格也添加居中对齐

        # 各落地时间的总计 - 起始列数调整为9
        col_num = 9
        for landing_time in landing_time_headers:
            total_volume = total_by_landing_time.get(landing_time, 0)
            cell = ws5.cell(row=total_row, column=col_num, value=total_volume if total_volume > 0 else None)
            cell.font = Font(bold=True)
            cell.fill = total_fill
            cell.border = thin_border
            col_num += 1

        # 总的总计
        grand_total_cell = ws5.cell(row=total_row, column=len(all_headers), value=grand_total)
        grand_total_cell.font = Font(bold=True)
        grand_total_cell.fill = total_fill
        grand_total_cell.border = thin_border

        # 为总计行的其他空白单元格设置样式
        for col in range(2, 9):  # 列2-8的空白单元格（因为增加了商机情况列）
            cell = ws5.cell(row=total_row, column=col)
            cell.fill = total_fill
            cell.border = thin_border

        self._auto_adjust_column_width(ws5)
        ws5.freeze_panes = 'E2'
         
        # ==================== 创建响应 ====================
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        current_time = timezone.now().strftime('%Y%m%d_%H%M%S')
        filename = f'综合调研数据工作簿_{current_time}.xlsx'
        encoded_filename = quote(filename.encode('utf-8'))
        response['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{encoded_filename}'
        
        wb.save(response)
        
        return response

    download_comprehensive_workbook.short_description = "导出综合excel"
    download_comprehensive_workbook.type = 'success'
    download_comprehensive_workbook.style = 'color:white;'

    # 辅助方法 - 为Sheet3提供数据分组
    def _group_all_hospital_data_for_sheet3(self, hospital_surveys_queryset):
        """为Sheet3提供数据分组"""
        from collections import defaultdict
        
        grouped_data = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
        
        for hospital_survey in hospital_surveys_queryset.select_related(
            'hospital',
            'qitian_manager',
            'director_contact',
            'leader_contact',
            'operator_contact',
            'created_by'
        ):
            # 确保计算最新统计
            hospital_survey.calculate_all_statistics()
            
            # 获取该医院的所有仪器调研
            instruments = MindrayInstrumentSurvey.objects.filter(
                hospital_survey=hospital_survey,
                is_active=True
            ).select_related(
                'category',
                'brand',
                'competitionrelation',
                'last_modified_by'
            )
            
            if not instruments.exists():
                # 如果医院没有仪器调研，仍然要显示医院信息
                grouped_data[hospital_survey] = {}
            else:
                # 按仪器分类组织数据
                for instrument in instruments:
                    category_name = instrument.category.name if instrument.category else '未分类'
                    
                    if instrument.category and instrument.category.name == '血球':
                        blood_projects = MindrayBloodCellProject.objects.filter(
                            instrument_survey=instrument,
                            is_active=True
                        ).select_related('competitionrelation')
                        
                        if blood_projects.exists():
                            grouped_data[hospital_survey][category_name][instrument].extend(blood_projects)
                        else:
                            grouped_data[hospital_survey][category_name][instrument].append(None)
                    else:
                        grouped_data[hospital_survey][category_name][instrument].append(None)
        
        return grouped_data

    def _write_data_row_for_sheet3(self, ws, row_num, hospital_survey, instrument, project, familiarity_map, thin_border):
        """为Sheet3写入一行数据"""
        
        # 计算标本量总和
        total_sample_volume = 0
        if instrument:
            if instrument.category and instrument.category.name == '血球':
                from django.db.models import Sum
                total_sample_volume = MindrayBloodCellProject.objects.filter(
                    instrument_survey=instrument,
                    is_active=True
                ).aggregate(total=Sum('sample_volume'))['total'] or 0
            else:
                total_sample_volume = instrument.sample_volume or 0
        
        # 统一竞品关系点显示
        competition_relation_display = ''
        if instrument:
            if instrument.category and instrument.category.name == '血球':
                competition_relation_display = instrument.blood_competition_relations or '-'
            else:
                competition_relation_display = instrument.competitionrelation.competitionrelation if instrument.competitionrelation else '-'
        
        # 统一经销商显示
        dealer_name_display = ''
        if instrument:
            if instrument.category and instrument.category.name == '血球':
                dealer_name_display = instrument.blood_dealer_names or '-'
            else:
                dealer_name_display = instrument.dealer_name or '-'
        
        # 构建完整行数据
        row_data = [
            # 医院调研信息（列1-27）
            hospital_survey.hospital.hospitalname if hospital_survey and hospital_survey.hospital else '',
            hospital_survey.hospital.district if hospital_survey and hospital_survey.hospital else '',
            hospital_survey.hospital.hospitalclass if hospital_survey and hospital_survey.hospital else '',
            hospital_survey.qitian_manager.chinesename if hospital_survey and hospital_survey.qitian_manager else '',
            hospital_survey.mindray_manager if hospital_survey else '',
            hospital_survey.director_name if hospital_survey else '',
            hospital_survey.director_contact.chinesename if hospital_survey and hospital_survey.director_contact else '',
            familiarity_map.get(hospital_survey.director_familiarity, hospital_survey.director_familiarity or '') if hospital_survey else '',
            hospital_survey.leader_name if hospital_survey else '',
            hospital_survey.leader_contact.chinesename if hospital_survey and hospital_survey.leader_contact else '',
            familiarity_map.get(hospital_survey.leader_familiarity, hospital_survey.leader_familiarity or '') if hospital_survey else '',
            hospital_survey.operator_name if hospital_survey else '',
            hospital_survey.operator_contact.chinesename if hospital_survey and hospital_survey.operator_contact else '',
            dict(MindrayHospitalSurvey.SALES_MODE_CHOICES).get(hospital_survey.sales_mode, hospital_survey.sales_mode or '') if hospital_survey else '',
            hospital_survey.distribution_channel if hospital_survey else '',
            hospital_survey.routine_total_volume if hospital_survey else 0,
            hospital_survey.crp_total_volume if hospital_survey else 0,
            hospital_survey.saa_total_volume if hospital_survey else 0,
            hospital_survey.esr_total_volume if hospital_survey else 0,
            hospital_survey.glycation_total_volume if hospital_survey else 0,
            hospital_survey.urine_total_volume if hospital_survey else 0,
            hospital_survey.blood_cell_total_count if hospital_survey else 0,
            hospital_survey.glycation_total_count if hospital_survey else 0,
            hospital_survey.urine_total_count if hospital_survey else 0,
            hospital_survey.blood_cell_summary if hospital_survey else '',
            hospital_survey.glycation_summary if hospital_survey else '',
            hospital_survey.urine_summary if hospital_survey else '',
            
            # 仪器分类信息（列28）
            instrument.category.name if instrument and instrument.category else '',
            
            # 仪器调研信息（列29-39）
            '是' if instrument and instrument.is_our_instrument else ('否' if instrument else ''),
            dict(MindrayInstrumentSurvey.SALES_CHANNEL_CHOICES).get(instrument.our_sales_channel, instrument.our_sales_channel or '') if instrument and instrument.our_sales_channel else '',
            instrument.brand.brand if instrument and instrument.brand else '',
            instrument.model if instrument else '',
            instrument.quantity if instrument else '',
            instrument.installation_year if instrument else '',
            dict(MindrayInstrumentSurvey.installation_location_CHOICES).get(instrument.installation_location, instrument.installation_location or '') if instrument and instrument.installation_location else '',
            total_sample_volume,
            instrument.blood_project_details or '' if instrument and instrument.category and instrument.category.name == '血球' else '',
            competition_relation_display,
            dealer_name_display,
            
            # 血球项目详情（列40-43）
            project.get_project_type_display() if project else '',
            project.sample_volume or 0 if project else '',
            project.competitionrelation.competitionrelation if project and project.competitionrelation else ('未知' if project else ''),
            project.dealer_name or '未知' if project and project.dealer_name else ('未知' if project else ''),
            
            # 时间信息（列44-45）
            hospital_survey.created_by.chinesename if hospital_survey and hospital_survey.created_by else '',
            self._format_datetime(hospital_survey.updatetime) if hospital_survey else ''
        ]
        
        # 写入数据并设置边框
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _merge_cells_for_sheet3(self, ws, merge_ranges, familiarity_color_map, familiarity_font_color_map):
        """为Sheet3执行单元格合并并设置样式"""
        for merge_range in merge_ranges:
            start_row = merge_range['start_row']
            end_row = merge_range['end_row']
            start_col = merge_range['start_col']
            end_col = merge_range['end_col']
            merge_type = merge_range['type']
            
            # 只有当需要合并的行数大于1时才执行合并
            if end_row > start_row:
                for col in range(start_col, end_col + 1):
                    # 在合并前先保存客情颜色信息
                    familiarity_info = None
                    if col == 8 or col == 11:  # 主任客情(第8列)或组长客情(第11列)
                        first_cell = ws.cell(row=start_row, column=col)
                        if first_cell.value in ['不认识', '有商机在跟进', '有明确代理商', '成单']:
                            # 根据值确定颜色key
                            value_to_key = {
                                '不认识': 'red',
                                '有商机在跟进': 'yellow', 
                                '有明确代理商': 'green',
                                '成单': 'blue'
                            }
                            familiarity_info = value_to_key.get(first_cell.value)
                    
                    # 合并单元格
                    ws.merge_cells(
                        start_row=start_row, 
                        start_column=col,
                        end_row=end_row, 
                        end_column=col
                    )
                    
                    # 设置合并单元格的样式
                    merged_cell = ws.cell(row=start_row, column=col)
                    merged_cell.alignment = Alignment(
                        horizontal="center", 
                        vertical="center",
                        wrap_text=True
                    )
                    
                    # 优先设置客情颜色，如果是客情列的话
                    if familiarity_info and (col == 8 or col == 11):
                        bg_color = familiarity_color_map.get(familiarity_info)
                        font_color = familiarity_font_color_map.get(familiarity_info)
                        
                        if bg_color and font_color:
                            merged_cell.fill = PatternFill(
                                start_color=bg_color, 
                                end_color=bg_color, 
                                fill_type="solid"
                            )
                            merged_cell.font = Font(color=font_color, bold=True)
                    else:
                        # 非客情列才设置默认的合并背景色
                        if merge_type == 'category':
                            merged_cell.fill = PatternFill(
                                start_color="F0F8E8", 
                                end_color="F0F8E8", 
                                fill_type="solid"
                            )
                        elif merge_type == 'instrument':
                            merged_cell.fill = PatternFill(
                                start_color="FFF8E8", 
                                end_color="FFF8E8", 
                                fill_type="solid"
                            )

    def _reapply_familiarity_colors_for_sheet3(self, ws, data_by_hospital, familiarity_map, familiarity_color_map, familiarity_font_color_map):
        """为Sheet3重新应用客情颜色"""
        current_row = 2
        
        def paint_familiarity_cell(r, c, familiarity_key):
            """为客情单元格设置颜色"""
            if not familiarity_key:
                return
                
            bg = familiarity_color_map.get(familiarity_key)
            fg = familiarity_font_color_map.get(familiarity_key)
            if not (bg and fg):
                return
                
            target_cell = ws.cell(row=r, column=c)
            
            # 设置背景色
            target_cell.fill = PatternFill(
                start_color=bg.upper(), 
                end_color=bg.upper(), 
                fill_type="solid"
            )
            
            # 设置字体颜色和粗体
            target_cell.font = Font(color=fg.upper(), bold=True)
            
            # 保持居中对齐
            target_cell.alignment = Alignment(
                horizontal="center", 
                vertical="center",
                wrap_text=True
            )
        
        for hospital_survey, categories_data in data_by_hospital.items():
            hospital_start_row = current_row
            
            if not categories_data:
                # 医院没有仪器调研的情况
                # 主任客情（第8列）
                if hospital_survey.director_familiarity:
                    paint_familiarity_cell(current_row, 8, hospital_survey.director_familiarity)
                
                # 组长客情（第11列）
                if hospital_survey.leader_familiarity:
                    paint_familiarity_cell(current_row, 11, hospital_survey.leader_familiarity)
                
                current_row += 1
            else:
                # 有仪器调研的情况
                total_rows_in_hospital = sum(
                    sum(len(projects) if projects else 1 for projects in instruments_data.values())
                    for instruments_data in categories_data.values()
                )
                
                # 为医院的所有行设置客情颜色（因为医院信息会被合并）
                for row_offset in range(total_rows_in_hospital):
                    target_row = hospital_start_row + row_offset
                    
                    # 主任客情（第8列）
                    if hospital_survey.director_familiarity:
                        paint_familiarity_cell(target_row, 8, hospital_survey.director_familiarity)
                    
                    # 组长客情（第11列） 
                    if hospital_survey.leader_familiarity:
                        paint_familiarity_cell(target_row, 11, hospital_survey.leader_familiarity)
                
                current_row += total_rows_in_hospital 

    def _format_datetime(self, dt):
        """格式化日期时间"""
        if dt:
            local_time = timezone.localtime(dt)
            return local_time.strftime('%Y-%m-%d %H:%M:%S')
        return ''

    def _auto_adjust_column_width(self, ws):
        """自动调整列宽"""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max(max_length + 2, 10), 60)
            ws.column_dimensions[column_letter].width = adjusted_width
    

   #刷新计算
    def refresh_all_calculated_fields(self, request, queryset):
        """批量刷新所有计算字段（不修改updatetime）"""
        updated_count = 0
        
        for hospital_survey in queryset:
            # 先刷新所有仪器的汇总信息
            instruments = MindrayInstrumentSurvey.objects.filter(
                hospital_survey=hospital_survey, 
                is_active=True
            )
            
            for instrument in instruments:
                if instrument.is_blood_category:
                    # 刷新血球仪器的汇总信息（不修改updatetime）
                    instrument.calculate_all_blood_summaries()
                    
                    # 使用 update() 避免触发 auto_now
                    MindrayInstrumentSurvey.objects.filter(id=instrument.id).update(
                        sample_volume=instrument.sample_volume,
                        blood_project_types=instrument.blood_project_types,
                        blood_project_details=instrument.blood_project_details,
                        blood_competition_relations=instrument.blood_competition_relations,
                        blood_dealer_names=instrument.blood_dealer_names
                    )
            
            # 然后刷新医院调研的所有统计信息（包括商机汇总，不修改updatetime）
            hospital_survey.calculate_all_statistics()
            
            # 使用 update() 避免触发 auto_now
            MindrayHospitalSurvey.objects.filter(id=hospital_survey.id).update(
                crp_total_volume=hospital_survey.crp_total_volume,
                saa_total_volume=hospital_survey.saa_total_volume,
                esr_total_volume=hospital_survey.esr_total_volume,
                routine_total_volume=hospital_survey.routine_total_volume,
                glycation_total_volume=hospital_survey.glycation_total_volume,
                urine_total_volume=hospital_survey.urine_total_volume,
                blood_cell_total_count=hospital_survey.blood_cell_total_count,
                glycation_total_count=hospital_survey.glycation_total_count,
                urine_total_count=hospital_survey.urine_total_count,
                blood_cell_summary=hospital_survey.blood_cell_summary,
                glycation_summary=hospital_survey.glycation_summary,
                urine_summary=hospital_survey.urine_summary,
                sales_opportunities_summary=hospital_survey.sales_opportunities_summary  # 新增：商机汇总
            )
            
            updated_count += 1
        
        self.message_user(
            request, 
            f'成功刷新了 {updated_count} 条医院调研记录的所有计算字段（包括商机汇总）！（未修改更新时间）'
        )

    refresh_all_calculated_fields.short_description = "刷新计算"
    refresh_all_calculated_fields.type = 'info'
    refresh_all_calculated_fields.style = 'color:white;'

    def changelist_view(self, request, extra_context=None):
        """自定义列表页面，添加统计信息和图表数据"""
        extra_context = extra_context or {}
        
        # 获取当前查询集
        queryset = self.get_queryset(request)
        
        # 应用搜索和过滤
        cl = self.get_changelist_instance(request)
        queryset = cl.get_queryset(request)
        
        from django.db.models import Sum, Count, Avg, Q
        from django.utils import timezone
        from datetime import timedelta
        
        # 1. 基本统计
        total_hospitals = queryset.count()
        
        # 2. 标本量统计汇总 - 确保所有字段都包含
        volume_stats = queryset.aggregate(
            total_routine=Sum('routine_total_volume'),
            total_crp=Sum('crp_total_volume'),
            total_saa=Sum('saa_total_volume'),
            total_esr=Sum('esr_total_volume'),
            total_glycation=Sum('glycation_total_volume'),
            total_urine=Sum('urine_total_volume'),
        )
        
        
        # 3. 仪器台数统计汇总
        instrument_stats = queryset.aggregate(
            total_blood_cell=Sum('blood_cell_total_count'),
            total_glycation_count=Sum('glycation_total_count'),
            total_urine_count=Sum('urine_total_count'),
        )
        
        # 4. 按医院级别统计
        class_stats = queryset.values(
            'hospital__hospitalclass'
        ).annotate(
            count=Count('id')
        ).exclude(
            hospital__hospitalclass__isnull=True
        ).order_by('-count')
        
        # 5. 客情统计
        director_familiarity_stats = queryset.values('director_familiarity').annotate(
            count=Count('id')
        )
        leader_familiarity_stats = queryset.values('leader_familiarity').annotate(
            count=Count('id')
        )
        
        familiarity_map = {
            'red': '不认识',
            'yellow': '有商机在跟进',
            'green': '有明确代理商',
            'blue': '成单'
        }
        
        director_familiarity_display = {}
        leader_familiarity_display = {}
        
        for item in director_familiarity_stats:
            key = item['director_familiarity']
            if key:
                director_familiarity_display[familiarity_map.get(key, key)] = item['count']
        
        for item in leader_familiarity_stats:
            key = item['leader_familiarity']
            if key:
                leader_familiarity_display[familiarity_map.get(key, key)] = item['count']
        
        # 6. 修改时间统计
        today = timezone.now().date()
        week_start = today - timedelta(days=today.weekday())
        month_start = today.replace(day=1)
        
        this_week_count = queryset.filter(updatetime__date__gte=week_start).count()
        this_month_count = queryset.filter(updatetime__date__gte=month_start).count()
        
        # 7. 医院标本量排名（Top15）
        hospital_sample_rankings = []
        for hospital in queryset.select_related('hospital'):
            total_sample = (
                (hospital.crp_total_volume or 0) +
                (hospital.saa_total_volume or 0) +
                (hospital.esr_total_volume or 0) +
                (hospital.routine_total_volume or 0) +
                (hospital.glycation_total_volume or 0) +
                (hospital.urine_total_volume or 0)
            )
            if total_sample > 0:
                hospital_sample_rankings.append({
                    'hospital_name': hospital.hospital.hospitalname,
                    'total_sample': total_sample,
                    'district': hospital.hospital.district,
                    'rank': 0  # 将在排序后设置
                })
        
        # 按总标本量排序并添加排名
        hospital_sample_rankings.sort(key=lambda x: x['total_sample'], reverse=True)
        for i, hospital in enumerate(hospital_sample_rankings[:15], 1):
            hospital['rank'] = i
        
        # 8. 医院仪器台数排名（Top15）
        hospital_instrument_rankings = []
        for hospital in queryset.select_related('hospital'):
            total_instruments = (
                (hospital.blood_cell_total_count or 0) +
                (hospital.glycation_total_count or 0) +
                (hospital.urine_total_count or 0)
            )
            if total_instruments > 0:
                hospital_instrument_rankings.append({
                    'hospital_name': hospital.hospital.hospitalname,
                    'total_instruments': total_instruments,
                    'district': hospital.hospital.district,
                    'rank': 0  # 将在排序后设置
                })
        
        # 按总仪器台数排序并添加排名
        hospital_instrument_rankings.sort(key=lambda x: x['total_instruments'], reverse=True)
        for i, hospital in enumerate(hospital_instrument_rankings[:15], 1):
            hospital['rank'] = i
        
        # 9. 修改：医院商机数量排名（基于SalesOpportunity数量）
        from .models import SalesOpportunity  # 确保导入SalesOpportunity
        
        hospital_opportunity_rankings = []
        for hospital in queryset.select_related('hospital'):
            # 统计该医院调研记录的商机数量
            opportunity_count = SalesOpportunity.objects.filter(
                hospital_survey=hospital,  # 使用正确的字段名
                is_active=True
            ).count()
            
            if opportunity_count > 0:
                hospital_opportunity_rankings.append({
                    'hospital_name': hospital.hospital.hospitalname,
                    'opportunity_count': opportunity_count,
                    'district': hospital.hospital.district,
                    'rank': 0  # 将在排序后设置
                })
        
        # 按商机数量排序并添加排名
        hospital_opportunity_rankings.sort(key=lambda x: x['opportunity_count'], reverse=True)
        for i, hospital in enumerate(hospital_opportunity_rankings[:15], 1):
            hospital['rank'] = i
        
        # 准备图表数据 - 修正标本量数据
        # 标本量分布数据（按值降序排序） - 确保包含所有类型
        volume_data_raw = [
            (volume_stats['total_routine'] or 0, '血常规'),
            (volume_stats['total_crp'] or 0, 'CRP'),
            (volume_stats['total_saa'] or 0, 'SAA'),  # 确保SAA包含
            (volume_stats['total_esr'] or 0, '血沉'),
            (volume_stats['total_glycation'] or 0, '糖化'),
            (volume_stats['total_urine'] or 0, '尿液'),  # 确保尿液包含
        ]

       


        # 过滤掉值为0的项目（避免图表中显示0值）
        volume_data_filtered = [(value, label) for value, label in volume_data_raw if value > 0]
        volume_data_filtered.sort(key=lambda x: x[0], reverse=True)
        
        volume_chart_data = [item[0] for item in volume_data_filtered]
        volume_chart_labels = [item[1] for item in volume_data_filtered]
        
        # 仪器台数分布数据（按值降序排序）
        instrument_data_raw = [
            (instrument_stats['total_blood_cell'] or 0, '血球仪器'),
            (instrument_stats['total_glycation_count'] or 0, '糖化仪器'),
            (instrument_stats['total_urine_count'] or 0, '尿液仪器'),
        ]
        # 过滤掉值为0的项目
        instrument_data_filtered = [(value, label) for value, label in instrument_data_raw if value > 0]
        instrument_data_filtered.sort(key=lambda x: x[0], reverse=True)
        
        instrument_chart_data = [item[0] for item in instrument_data_filtered]
        instrument_chart_labels = [item[1] for item in instrument_data_filtered]
        
        extra_context.update({
            # 基本统计
            'total_hospitals': total_hospitals,
            'this_week_count': this_week_count,
            'this_month_count': this_month_count,
            'total_all_samples': sum(volume_chart_data) if volume_chart_data else 0,
            'total_all_instruments': sum(instrument_chart_data) if instrument_chart_data else 0,
            
            # 图表数据
            'volume_chart_data': volume_chart_data,
            'volume_chart_labels': volume_chart_labels,
            'instrument_chart_data': instrument_chart_data,
            'instrument_chart_labels': instrument_chart_labels,
            
            # 传统统计保留
            'volume_stats': volume_stats,
            'instrument_stats': instrument_stats,
            'class_stats': class_stats,
            'director_familiarity_stats': director_familiarity_display,
            'leader_familiarity_stats': leader_familiarity_display,
            
            # 排名数据
            'hospital_sample_rankings': hospital_sample_rankings[:15],
            'hospital_instrument_rankings': hospital_instrument_rankings[:15],
            'hospital_opportunity_rankings': hospital_opportunity_rankings[:15],
        })
        
        #    # 添加这个调试，看看传递给模板的数据
        # print(f"传递给模板的 volume_chart_data: {extra_context['volume_chart_data']}")
        # print(f"传递给模板的 volume_chart_labels: {extra_context['volume_chart_labels']}")
        # print("=== 调试结束 ===")
        
        return super().changelist_view(request, extra_context=extra_context)    


    # 权限控制方法
    def has_add_permission(self, request):
        """只有管理员可以新增医院调研"""
        return request.user.is_superuser
    
    def has_delete_permission(self, request, obj=None):
        """只有管理员可以删除医院调研"""
        return request.user.is_superuser
    
    def has_change_permission(self, request, obj=None):
        """销售只能修改自己负责的医院调研"""
        if request.user.is_superuser:
            return True
            
        if not request.user.username in ALLOWED_SALESMEN:
            return False
            
        # 如果是查看列表页面，允许访问
        if obj is None:
            return True
            
        # 如果是修改具体对象，检查是否是自己负责的医院
        return obj.qitian_manager == request.user

    def get_readonly_fields(self, request, obj=None):
        """根据用户权限设置只读字段"""
        readonly_fields = list(super().get_readonly_fields(request, obj))
        
        # 如果不是管理员，且不是自己负责的医院，设置基础信息为只读
        if not request.user.is_superuser:
            if obj and obj.qitian_manager != request.user:
                # 不是自己负责的医院，所有字段都只读
                readonly_fields.extend([
                    'hospital', 'qitian_manager', 'mindray_manager',
                    'director_name', 'director_contact', 'director_familiarity',
                    'leader_name', 'leader_contact', 'leader_familiarity',
                    'operator_name', 'operator_contact',
                    'sales_mode', 'distribution_channel'
                ])
        
        return readonly_fields


    def get_inline_instances(self, request, obj=None):
        """根据权限控制inline显示"""
        inline_instances = []
        
        # 管理员可以看到所有inline
        if request.user.is_superuser:
            return super().get_inline_instances(request, obj)
        
        # 销售只能在自己负责的医院中编辑仪器和商机
        if obj and obj.qitian_manager == request.user:
            return super().get_inline_instances(request, obj)
        
        # 其他情况只显示只读的inline
        if obj:
            # 创建只读的inline实例
            inline_instances = super().get_inline_instances(request, obj)
            for inline in inline_instances:
                inline.has_add_permission = lambda r, o=None: False
                inline.has_change_permission = lambda r, o=None: False
                inline.has_delete_permission = lambda r, o=None: False
        
        return inline_instances


    def delete_model(self, request, obj):
        """删除单个医院调研时级联删除所有关联数据"""
        print('im in hospital survey delete_model')
        
        # 1. 先逻辑删除所有关联的血球项目
        MindrayBloodCellProject.objects.filter(
            instrument_survey__hospital_survey=obj,
            is_active=True
        ).update(is_active=False)
        
        # 2. 逻辑删除所有关联的仪器调研
        MindrayInstrumentSurvey.objects.filter(
            hospital_survey=obj,
            is_active=True
        ).update(is_active=False)
        
        # 3. 逻辑删除所有关联的销售商机
        SalesOpportunity.objects.filter(
            hospital_survey=obj,
            is_active=True
        ).update(is_active=False)
        
        # 4. 最后删除医院调研本身
        obj.is_active = False 
        obj.save()

    def delete_queryset(self, request, queryset):
        """批量删除医院调研时级联删除所有关联数据"""
        print('im in hospital survey delete_queryset')
        
        for obj in queryset:
            # 1. 先逻辑删除所有关联的血球项目
            MindrayBloodCellProject.objects.filter(
                instrument_survey__hospital_survey=obj,
                is_active=True
            ).update(is_active=False)
            
            # 2. 逻辑删除所有关联的仪器调研
            MindrayInstrumentSurvey.objects.filter(
                hospital_survey=obj,
                is_active=True
            ).update(is_active=False)
            
            # 3. 逻辑删除所有关联的销售商机
            SalesOpportunity.objects.filter(
                hospital_survey=obj,
                is_active=True
            ).update(is_active=False)
        
        # 4. 最后批量删除医院调研
        queryset.update(is_active=False)

    #让销售在仪器界面和商机界面选择医院的时候只能选自己负责 
    def get_search_results(self, request, queryset, search_term):
        """处理 autocomplete 搜索的权限控制"""
        queryset, use_distinct = super().get_search_results(request, queryset, search_term)
        
        # 如果是 autocomplete 请求
        if 'autocomplete' in request.path:
            print(f"MindrayHospitalSurvey autocomplete 请求，当前用户: {request.user.username}")
            
            # 如果是普通销售，只能选择自己负责的医院调研
            if not request.user.is_superuser and request.user.username in ALLOWED_SALESMEN:
                print("在 MindrayHospitalSurvey 中应用权限过滤")
                queryset = queryset.filter(qitian_manager=request.user)
                print(f"过滤后数量: {queryset.count()}")
        
        return queryset, use_distinct

    class Media:
        css = {
            'all': ('admin/css/freeze_firsttwo_column.css',)
        }



@admin.register(MindrayInstrumentSurvey)
class MindrayInstrumentSurveyAdmin(GlobalAdmin):
    ordering = ['hospital_survey__id','category','brand','createtime']
    form = MindrayInstrumentSurveyForm
    autocomplete_fields = ['hospital_survey', 'brand', 'competitionrelation']
    
    
    list_display = [
        'hospital_survey', 
        'get_hospital_district',
        'get_hospital_class',
        'get_hospital_manager',
        'get_category_display', 
        'is_our_instrument', 
        'our_sales_channel',
        'brand', 
        'model', 
        'quantity', 
        'installation_year', 
        'installation_location',
        'get_total_sample_volume',
        'get_blood_project_details_display',     # 新增：血球项目详细汇总
        'get_competition_relation_display',      # 新增：统一竞品关系显示
        'get_dealer_name_display',               # 新增：统一经销商显示
        'get_last_modified_by',
        'updatetime'
    ]

    list_filter = [
        'category', 
        'is_our_instrument', 
        'our_sales_channel',
        'brand', 
        HospitalManagerFilter,  # 1. 添加医院负责人筛选
        InstallationYearFilter,  # 2. 修正的装机年份筛选
        'hospital_survey__hospital__district',
        'hospital_survey__hospital__hospitalclass',
    ]
    
    search_fields = [
        'hospital_survey__hospital__hospitalname', 
        'model', 
        'installation_location', 
        'dealer_name',
        'hospital_survey__qitian_manager__chinesename',
        'last_modified_by__chinesename',  # 新增：支持按修改人搜索
    ]
    
    fieldsets = (
        ('基本信息', {
            'fields': (
                ('hospital_survey', 'category'),
                ('is_our_instrument', 'our_sales_channel'),
                ('brand', 'model'),
                ('quantity', 'installation_year'),
                ('sample_volume', 'installation_location'),
                ('competitionrelation', 'dealer_name'),
            ),
            'classes': ('primary-fieldset',)
        }),
    )
    

    # # 让仪器分类有颜色   
    def get_category_display(self, obj):
        """显示带颜色和背景的仪器分类"""
        if not obj.category:
            return '-'
        
        category_name = obj.category.name
        
        # 定义样式映射
        style_mapping = {
            '血球': {
                'color': '#FFFFFF',
                'background': '#FF6B6B',
                'border': '1px solid #FF6B6B'
            },
            '糖化': {
                'color': '#FFFFFF', 
                'background': '#F5A512',
                'border': '1px solid #F5A512'
            },
            '尿液': {
                'color': '#FFFFFF',
                'background': '#45B7D1', 
                'border': '1px solid #45B7D1'
            }
        }
        
        # 获取对应样式，默认样式
        styles = style_mapping.get(category_name, {
            'color': '#333333',
            'background': '#F8F9FA',
            'border': '1px solid #DEE2E6'
        })
        
        style_str = f"color: {styles['color']}; background-color: {styles['background']}; border: {styles['border']}; padding: 3px 8px; border-radius: 4px; font-weight: bold; font-size: 12px;"
        
        return format_html(
            '<span style="{}">{}</span>',
            style_str,
            category_name
        )
 
    get_category_display.short_description = '仪器分类'
    get_category_display.admin_order_field = 'category__name'


    def get_hospital_district(self, obj):
        if obj.hospital_survey and obj.hospital_survey.hospital:
            return obj.hospital_survey.hospital.district
        return '-'
    get_hospital_district.short_description = '区域'
    get_hospital_district.admin_order_field = 'hospital_survey__hospital__district'
    
    def get_hospital_class(self, obj):
        if obj.hospital_survey and obj.hospital_survey.hospital:
            return obj.hospital_survey.hospital.hospitalclass
        return '-'
    get_hospital_class.short_description = '级别'
    get_hospital_class.admin_order_field = 'hospital_survey__hospital__hospitalclass'
    

    # 新增：最后修改人显示方法
    def get_last_modified_by(self, obj):
        if hasattr(obj, 'last_modified_by') and obj.last_modified_by:
            return obj.last_modified_by.chinesename or obj.last_modified_by.username
        return '-'
    get_last_modified_by.short_description = '最后修改人'
    get_last_modified_by.admin_order_field = 'last_modified_by__chinesename'
    
    # 3. 添加医院负责人显示方法
    def get_hospital_manager(self, obj):
        if obj.hospital_survey and obj.hospital_survey.qitian_manager:
            return obj.hospital_survey.qitian_manager.chinesename
        return '-'
    get_hospital_manager.short_description = '医院负责人'
    get_hospital_manager.admin_order_field = 'hospital_survey__qitian_manager__chinesename'
    
    # 4. 统一的标本量总和计算显示
    def get_total_sample_volume(self, obj):
        """
        根据仪器类别计算标本量总和：
        - 血球仪器：计算所有血球项目的标本量总和
        - 糖化/尿液仪器：直接返回sample_volume字段值
        """
        if obj.category and obj.category.name == '血球':
            # 血球仪器：计算该仪器下所有血球项目的标本量总和
            from django.db.models import Sum
            total_volume = MindrayBloodCellProject.objects.filter(
                instrument_survey=obj,
                is_active=True
            ).aggregate(total=Sum('sample_volume'))['total'] or 0
            return total_volume
        else:
            # 糖化/尿液仪器：返回sample_volume字段值
            return obj.sample_volume or 0
    
    get_total_sample_volume.short_description = '标本量总和'
    get_total_sample_volume.admin_order_field = 'sample_volume'  # 默认按sample_volume排序
    


    # 新增显示方法
    def get_blood_projects_display(self, obj):
        """显示血球项目汇总"""
        if obj.category and obj.category.name == '血球':
            return obj.blood_project_types or '-'
        return '-'
    get_blood_projects_display.short_description = '血球项目'
    get_blood_projects_display.admin_order_field = 'blood_project_types'
    
    def get_blood_competition_display(self, obj):
        """显示血球竞品关系点汇总，长文本截断"""
        if obj.category and obj.category.name == '血球':
            relations = obj.blood_competition_relations or '-'
            if len(relations) > 50:
                return f"{relations[:50]}..."
            return relations
        return '-'
    get_blood_competition_display.short_description = '血球竞品关系点'
    get_blood_competition_display.admin_order_field = 'blood_competition_relations'
    
    def get_blood_dealers_display(self, obj):
        """显示血球经销商汇总，长文本截断"""
        if obj.category and obj.category.name == '血球':
            dealers = obj.blood_dealer_names or '-'
            if len(dealers) > 50:
                return f"{dealers[:50]}..."
            return dealers
        return '-'
    get_blood_dealers_display.short_description = '血球经销商'
    get_blood_dealers_display.admin_order_field = 'blood_dealer_names'


    # # 新增：血球项目详细汇总显示
    # def get_blood_project_details_display(self, obj):
    #     """显示血球项目详细汇总，只针对血球仪器"""
    #     if obj.category and obj.category.name == '血球':
    #         details = obj.blood_project_details or '-'
    #         # 如果内容太长，截断显示
    #         if len(details) > 80:
    #             return f"{details[:80]}..."
    #         return details
    #     return '-'
    # get_blood_project_details_display.short_description = '血球项目-标本量-竞品关系点-经销商'
    # get_blood_project_details_display.admin_order_field = 'blood_project_details'

    #============
    # 新增：血球项目详细汇总显示
    def get_blood_project_details_display(self, obj):
        """显示血球项目详细汇总，只针对血球仪器"""
        if obj.category and obj.category.name == '血球':
            details = obj.blood_project_details or '-'
            if details != '-':
                # 处理换行显示和截断逻辑
                if len(details) > 80:
                    # 截断时保持换行格式
                    truncated = f"{details[:80]}..."
                    formatted_details = truncated.replace('\n', '<br>')
                else:
                    formatted_details = details.replace('\n', '<br>')
                return format_html(formatted_details)
            return details
        return '-'
    get_blood_project_details_display.short_description = '血球项目-标本量-竞品关系点-经销商'
    get_blood_project_details_display.admin_order_field = 'blood_project_details'
    #===========


    # 新增：统一竞品关系显示
    def get_competition_relation_display(self, obj):
        """
        统一竞品关系显示：
        - 血球仪器：显示blood_competition_relations
        - 糖化/尿液仪器：显示competitionrelation
        """
        if obj.category and obj.category.name == '血球':
            relations = obj.blood_competition_relations or '-'
            # 如果内容太长，截断显示
            if len(relations) > 50:
                return f"{relations[:50]}..."
            return relations
        else:
            # 糖化/尿液仪器
            if obj.competitionrelation:
                return obj.competitionrelation.competitionrelation
            return '-'
    get_competition_relation_display.short_description = '竞品关系点'
    get_competition_relation_display.admin_order_field = 'competitionrelation'

    # 新增：统一经销商显示
    def get_dealer_name_display(self, obj):
        """
        统一经销商显示：
        - 血球仪器：显示blood_dealer_names
        - 糖化/尿液仪器：显示dealer_name
        """
        if obj.category and obj.category.name == '血球':
            dealers = obj.blood_dealer_names or '-'
            # 如果内容太长，截断显示
            if len(dealers) > 50:
                return f"{dealers[:50]}..."
            return dealers
        else:
            # 糖化/尿液仪器
            return obj.dealer_name or '-'
    get_dealer_name_display.short_description = '经销商名称'
    get_dealer_name_display.admin_order_field = 'dealer_name'



    # 5. 动态添加inline，只对血球仪器显示项目详情
    def get_inlines(self, request, obj):
        inlines = []
        if obj and obj.category and obj.category.name == '血球':
            inlines.append(BloodCellProjectInlineForInstrument)
        return inlines    
    

    def get_queryset(self, request):
        return super().get_queryset(request).filter(is_active=True).select_related(
            'hospital_survey__qitian_manager',
            'hospital_survey__hospital',
            'category',
            'brand',
            'last_modified_by'  # 新增：优化查询性能
        )
     
    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        """根据权限限制外键选择范围"""
        if db_field.name == "hospital_survey":
            kwargs["queryset"] = MindrayHospitalSurvey.objects.filter(is_active=True)
            
            # 销售只能选择自己负责的医院调研
            if not request.user.is_superuser and request.user.username in ALLOWED_SALESMEN:
                kwargs["queryset"] = kwargs["queryset"].filter(qitian_manager=request.user)
        
        elif db_field.name == "category":
            kwargs["queryset"] = MindrayInstrumentCategory.objects.filter(is_active=True).order_by('order')
        elif db_field.name == "brand":
            kwargs["queryset"] = Brand.objects.filter(is_active=True)
        elif db_field.name == "competitionrelation":
            kwargs["queryset"] = CompetitionRelation.objects.filter(is_active=True)
        elif db_field.name == "last_modified_by":
            kwargs["queryset"] = UserInfoMindray.objects.filter(is_active=True)
        
        return super().formfield_for_foreignkey(db_field, request, **kwargs)
    
     
    def change_view(self, request, object_id, form_url='', extra_context=None):
        extra_context = extra_context or {}
        
        try:
            obj = self.get_object(request, object_id)
            if obj and obj.category and obj.category.name == '血球':
                # 获取血球项目详情
                projects = MindrayBloodCellProject.objects.filter(
                    instrument_survey=obj,
                    is_active=True
                )
                extra_context['blood_cell_projects'] = projects
                extra_context['is_blood_cell'] = True
                
                # 实时计算当前血球项目汇总信息
                obj.calculate_all_blood_summaries()
                extra_context['blood_cell_total_volume'] = obj.sample_volume
                extra_context['blood_project_types'] = obj.blood_project_types
                extra_context['blood_competition_relations'] = obj.blood_competition_relations
                extra_context['blood_dealer_names'] = obj.blood_dealer_names
            else:
                extra_context['is_blood_cell'] = False
                extra_context['current_sample_volume'] = obj.sample_volume or 0
                
            # 显示最后修改人信息
            if obj and hasattr(obj, 'last_modified_by') and obj.last_modified_by:
                extra_context['last_modified_by'] = obj.last_modified_by
                
        except:
            pass
            
        return super().change_view(request, object_id, form_url, extra_context)

    
    def get_fieldsets(self, request, obj=None):
        if obj and obj.category and obj.category.name == '血球':
            # 血球仪器的字段集
            return (
                ('基本信息', {
                    'fields': (
                        ('hospital_survey', 'category'),
                        ('is_our_instrument', 'our_sales_channel'),
                        ('brand', 'model'),
                        ('quantity', 'installation_year'),
                        'installation_location',
                    ),
                    'classes': ('primary-fieldset',)
                }),
                ('血球项目汇总信息 (自动计算)', {
                    'fields': (
                        'sample_volume',
                        'blood_project_types',
                        'blood_project_details',        # 新增显示
                        'blood_competition_relations',
                        'blood_dealer_names',
                    ),
                    'classes': ('secondary-fieldset', 'readonly-fieldset'),
                    'description': '以下字段由血球项目详情自动计算生成，请通过下方的"血球项目详情"添加或修改具体项目信息。'
                }),
            )
        else:
            # 非血球仪器的字段集（糖化/尿液）
            return (
                ('基本信息', {
                    'fields': (
                        ('hospital_survey', 'category'),
                        ('is_our_instrument', 'our_sales_channel'),
                        ('brand', 'model'),
                        ('quantity', 'installation_year'),
                        ('sample_volume', 'installation_location'),
                        ('competitionrelation', 'dealer_name'),
                    ),
                    'classes': ('primary-fieldset',)
                }),
            )
    
    def get_readonly_fields(self, request, obj=None):
        readonly_fields = list(super().get_readonly_fields(request, obj))
        
        # 为血球仪器的汇总字段设为只读
        if obj and obj.category and obj.category.name == '血球':
            blood_readonly_fields = [
                'sample_volume', 'blood_project_types', 
                'blood_project_details',           # 新增只读字段
                'blood_competition_relations', 'blood_dealer_names'
            ]
            for field in blood_readonly_fields:
                if field not in readonly_fields:
                    readonly_fields.append(field)
        
        # 如果有最后修改人字段，设为只读
        if hasattr(self.model, 'last_modified_by') and 'last_modified_by' not in readonly_fields:
            readonly_fields.append('last_modified_by')
        
        return readonly_fields

   
    def save_model(self, request, obj, form, change):
        """只有在实际有改动时才更新时间和触发联动更新"""
        # 设置最后修改人
        if hasattr(obj, 'last_modified_by'):
            obj.last_modified_by = request.user
        
        # 检查是否有实际改动
        has_changes = False
        if not change:  # 新建记录
            has_changes = True
        elif form.changed_data:  # 修改且有字段变化
            has_changes = True
        
        if not has_changes and change:
            # 没有实际改动，只刷新数据不保存
            obj.refresh_from_db()
            return
            
        obj.full_clean()
        super().save_model(request, obj, form, change)
        
        # 只有在有实际改动时才触发医院调研的统计计算
        if has_changes and obj.hospital_survey:
            obj.hospital_survey.calculate_all_statistics()
            
            # 手动更新hospital_survey的updatetime
            from django.utils import timezone
            obj.hospital_survey.updatetime = timezone.now()
            
            obj.hospital_survey.save(update_fields=[
                'crp_total_volume', 'saa_total_volume', 
                'esr_total_volume', 'routine_total_volume',
                'glycation_total_volume', 'urine_total_volume',
                'blood_cell_total_count', 'glycation_total_count', 'urine_total_count',
                'blood_cell_summary', 'glycation_summary', 'urine_summary',
                'updatetime'
            ])
    
    def save_formset(self, request, form, formset, change):
        """只有在实际有改动时才更新时间和统计"""
        # 检查是否是血球项目的formset
        if hasattr(formset, 'model') and formset.model == MindrayBloodCellProject:
            instances = formset.save(commit=False)
            deleted_objects = formset.deleted_objects
            
            # 检查是否有实际改动
            has_changes = bool(instances) or bool(deleted_objects)
            
            if not has_changes:
                return  # 没有改动，直接返回
            
            for instance in instances:
                instance.instrument_survey = form.instance
                instance.save()
            
            for obj in deleted_objects:
                obj.is_active = False
                obj.save()
            
            formset.save_m2m()
            
            # 重新计算并更新仪器的所有汇总信息
            form.instance.calculate_all_blood_summaries()
            form.instance.save(update_fields=[
                'sample_volume', 'blood_project_types', 
                'blood_project_details',
                'blood_competition_relations', 'blood_dealer_names'
                # updatetime由auto_now自动处理
            ])
            
            # 触发医院调研统计计算
            if form.instance.hospital_survey:
                from django.utils import timezone
                form.instance.hospital_survey.calculate_all_statistics()
                form.instance.hospital_survey.updatetime = timezone.now()
                form.instance.hospital_survey.save(update_fields=[
                    'crp_total_volume', 'saa_total_volume', 
                    'esr_total_volume', 'routine_total_volume',
                    'glycation_total_volume', 'urine_total_volume',
                    'blood_cell_total_count', 'glycation_total_count', 'urine_total_count',
                    'blood_cell_summary', 'glycation_summary', 'urine_summary',
                    'updatetime'
                ])
        else:
            # 非血球项目formset，使用默认处理
            super().save_formset(request, form, formset, change)

  
    def delete_model(self, request, obj):
        """删除单个仪器时级联删除血球项目并同步更新hospital_survey的统计信息"""
        # 在删除前先保存hospital_survey引用
        hospital_survey = obj.hospital_survey
        
        # 1. 先逻辑删除关联的血球项目
        MindrayBloodCellProject.objects.filter(
            instrument_survey=obj,
            is_active=True
        ).update(is_active=False)
        
        # 2. 如果是血球仪器，先清空相关汇总字段
        if obj.is_blood_category:
            obj.blood_project_details = ""
            obj.blood_project_types = ""
            obj.blood_competition_relations = ""
            obj.blood_dealer_names = ""
            obj.sample_volume = 0
        
        # 3. 执行删除操作（逻辑删除，设置is_active=False）
        obj.is_active = False
        obj.save()
        
        # 4. 删除后重新计算并更新hospital_survey统计
        if hospital_survey:
            from django.utils import timezone
            hospital_survey.calculate_all_statistics()
            hospital_survey.updatetime = timezone.now()
            hospital_survey.save(update_fields=[
                'crp_total_volume', 'saa_total_volume', 
                'esr_total_volume', 'routine_total_volume',
                'glycation_total_volume', 'urine_total_volume',
                'blood_cell_total_count', 'glycation_total_count', 'urine_total_count',
                'blood_cell_summary', 'glycation_summary', 'urine_summary',
                'sales_opportunities_summary',
                'updatetime'
            ])

    def delete_queryset(self, request, queryset):
        """批量删除仪器时级联删除血球项目并同步更新hospital_survey的统计信息"""
        # 在删除前收集所有相关的hospital_survey
        hospital_surveys = set()
        for obj in queryset:
            if obj.hospital_survey:
                hospital_surveys.add(obj.hospital_survey)
            
            # 1. 先逻辑删除关联的血球项目
            MindrayBloodCellProject.objects.filter(
                instrument_survey=obj,
                is_active=True
            ).update(is_active=False)
            
            # 2. 如果是血球仪器，先清空相关汇总字段
            if obj.is_blood_category:
                obj.blood_project_details = ""
                obj.blood_project_types = ""
                obj.blood_competition_relations = ""
                obj.blood_dealer_names = ""
                obj.sample_volume = 0
                obj.save(update_fields=[
                    'blood_project_details', 'blood_project_types',
                    'blood_competition_relations', 'blood_dealer_names', 'sample_volume'
                ])
        
        # 3. 执行批量删除操作
        queryset.update(is_active=False)
        
        # 4. 删除后重新计算并更新所有相关hospital_survey的统计
        from django.utils import timezone
        for hospital_survey in hospital_surveys:
            hospital_survey.calculate_all_statistics()
            hospital_survey.updatetime = timezone.now()
            hospital_survey.save(update_fields=[
                'crp_total_volume', 'saa_total_volume', 
                'esr_total_volume', 'routine_total_volume',
                'glycation_total_volume', 'urine_total_volume',
                'blood_cell_total_count', 'glycation_total_count', 'urine_total_count',
                'blood_cell_summary', 'glycation_summary', 'urine_summary',
                'sales_opportunities_summary',
                'updatetime'
            ])


    actions = ['download_simple_excel','download_comprehensive_excel','refresh_instrument_calculated_fields'] 

    def refresh_instrument_calculated_fields(self, request, queryset):
        """批量刷新仪器的计算字段（不修改updatetime）"""
        updated_count = 0
        hospital_surveys_to_update = set()
        
        for instrument in queryset:
            if instrument.is_blood_category:
                # 刷新血球仪器的汇总信息（不修改updatetime）
                instrument.calculate_all_blood_summaries()
                
                # 使用 update() 方法避免触发 auto_now
                MindrayInstrumentSurvey.objects.filter(id=instrument.id).update(
                    sample_volume=instrument.sample_volume,
                    blood_project_types=instrument.blood_project_types,
                    blood_project_details=instrument.blood_project_details,
                    blood_competition_relations=instrument.blood_competition_relations,
                    blood_dealer_names=instrument.blood_dealer_names
                )
                
                # 刷新实例以获取最新数据（但不修改updatetime）
                instrument.refresh_from_db()
            
            # 收集需要更新的医院调研
            if instrument.hospital_survey:
                hospital_surveys_to_update.add(instrument.hospital_survey)
            
            updated_count += 1
        
        # 更新相关的医院调研统计（不修改updatetime）
        for hospital_survey in hospital_surveys_to_update:
            hospital_survey.calculate_all_statistics()
            
            # 使用 update() 方法避免触发 auto_now
            MindrayHospitalSurvey.objects.filter(id=hospital_survey.id).update(
                crp_total_volume=hospital_survey.crp_total_volume,
                saa_total_volume=hospital_survey.saa_total_volume,
                esr_total_volume=hospital_survey.esr_total_volume,
                routine_total_volume=hospital_survey.routine_total_volume,
                glycation_total_volume=hospital_survey.glycation_total_volume,
                urine_total_volume=hospital_survey.urine_total_volume,
                blood_cell_total_count=hospital_survey.blood_cell_total_count,
                glycation_total_count=hospital_survey.glycation_total_count,
                urine_total_count=hospital_survey.urine_total_count,
                blood_cell_summary=hospital_survey.blood_cell_summary,
                glycation_summary=hospital_survey.glycation_summary,
                urine_summary=hospital_survey.urine_summary
            )
            
            # 刷新实例以获取最新数据
            hospital_survey.refresh_from_db()
        
        self.message_user(
            request, 
            f'成功刷新了 {updated_count} 条仪器记录的计算字段，同时更新了 {len(hospital_surveys_to_update)} 条医院调研统计！（未修改更新时间）'
        )    

    refresh_instrument_calculated_fields.short_description = "刷新计算字段"
    refresh_instrument_calculated_fields.type = 'info'
    refresh_instrument_calculated_fields.style = 'color:white;'

    refresh_instrument_calculated_fields.short_description = "刷新计算"

    refresh_instrument_calculated_fields.type = 'info'
    refresh_instrument_calculated_fields.style = 'color:white;'


    def _format_datetime(self, dt):
        """格式化日期时间"""
        if dt:
            local_time = timezone.localtime(dt)
            return local_time.strftime('%Y-%m-%d %H:%M:%S')
        return ''

    def _auto_adjust_column_width(self, ws):
        """自动调整列宽"""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max(max_length + 2, 10), 60)
            ws.column_dimensions[column_letter].width = adjusted_width
 
    def changelist_view(self, request, extra_context=None):
        """自定义列表页面，添加统计信息"""
        extra_context = extra_context or {}
        
        # 获取当前查询集（应用了筛选条件后的结果）
        queryset = self.get_queryset(request)
        
        # 应用搜索和过滤
        cl = self.get_changelist_instance(request)
        queryset = cl.get_queryset(request)
        
        from django.db.models import Sum, Count, Avg
        from django.utils import timezone
        from datetime import timedelta
        
        # 1. 基本统计
        total_instruments = queryset.count()
        total_quantity = queryset.aggregate(total=Sum('quantity'))['total'] or 0
        
        # 2. 按分类分别计算标本量
        blood_instruments = queryset.filter(category__name='血球')
        glycation_instruments = queryset.filter(category__name='糖化')
        urine_instruments = queryset.filter(category__name='尿液')
        
        # 血球仪器标本量（通过血球项目计算）
        blood_sample_volume = 0
        for instrument in blood_instruments:
            if instrument.category and instrument.category.name == '血球':
                from django.db.models import Sum
                total_volume = MindrayBloodCellProject.objects.filter(
                    instrument_survey=instrument,
                    is_active=True
                ).aggregate(total=Sum('sample_volume'))['total'] or 0
                blood_sample_volume += total_volume
            else:
                blood_sample_volume += instrument.sample_volume or 0
        
        # 糖化仪器标本量（直接取sample_volume字段）
        glycation_sample_volume = glycation_instruments.aggregate(
            total=Sum('sample_volume')
        )['total'] or 0
        
        # 尿液仪器标本量（直接取sample_volume字段）
        urine_sample_volume = urine_instruments.aggregate(
            total=Sum('sample_volume')
        )['total'] or 0
        
        total_sample_volume = blood_sample_volume + glycation_sample_volume + urine_sample_volume
        
        # 3. 按分类统计（按指定顺序）
        category_stats = {}
        category_order = ['血球', '糖化', '尿液']  # 指定顺序
        
        for category_name in category_order:
            cat_data = queryset.filter(category__name=category_name).aggregate(
                count=Count('id'),
                total_quantity=Sum('quantity')
            )
            if cat_data['count'] and cat_data['count'] > 0:
                category_stats[category_name] = {
                    'count': cat_data['count'],
                    'quantity': cat_data['total_quantity'] or 0
                }
        
        # 4. 按品牌统计（Top 5）
        brand_stats = queryset.exclude(brand__isnull=True).values(
            'brand__brand'
        ).annotate(
            count=Count('id'),
            total_quantity=Sum('quantity')
        ).order_by('-count')[:5]
        
        # 5. 按区域统计
        district_stats = queryset.values(
            'hospital_survey__hospital__district'
        ).annotate(
            count=Count('id'),
            total_quantity=Sum('quantity')
        ).order_by('-count')
        
        # 6. 我司vs非我司统计
        our_vs_competitor = queryset.values('is_our_instrument').annotate(
            count=Count('id'),
            total_quantity=Sum('quantity')
        )
        
        our_stats = {'count': 0, 'quantity': 0}
        competitor_stats = {'count': 0, 'quantity': 0}
        
        for item in our_vs_competitor:
            if item['is_our_instrument']:
                our_stats = {'count': item['count'], 'quantity': item['total_quantity'] or 0}
            else:
                competitor_stats = {'count': item['count'], 'quantity': item['total_quantity'] or 0}
        
        # 7. 装机年份分布（最近5年）
        current_year = timezone.now().year
        recent_years = [str(year) for year in range(current_year-4, current_year+1)]
        
        year_stats = {}
        for year in recent_years:
            year_count = queryset.filter(installation_year=year).count()
            if year_count > 0:
                year_stats[year] = year_count
        
        # 8. 本周/本月更新统计（改为updatetime）
        today = timezone.now().date()
        week_start = today - timedelta(days=today.weekday())
        month_start = today.replace(day=1)
        
        this_week_count = queryset.filter(updatetime__date__gte=week_start).count()
        this_month_count = queryset.filter(updatetime__date__gte=month_start).count()
        
        extra_context.update({
            # 基本统计
            'total_instruments': total_instruments,
            'total_quantity': total_quantity,
            'total_sample_volume': total_sample_volume,
            'blood_sample_volume': blood_sample_volume,
            'glycation_sample_volume': glycation_sample_volume,
            'urine_sample_volume': urine_sample_volume,
            
            # 分类统计
            'category_stats': category_stats,
            'brand_stats': brand_stats,
            'district_stats': district_stats,
            
            # 我司vs竞品
            'our_stats': our_stats,
            'competitor_stats': competitor_stats,
            'our_percentage': round(our_stats['count'] / total_instruments * 100, 1) if total_instruments > 0 else 0,
            
            # 时间统计
            'year_stats': year_stats,
            'this_week_count': this_week_count,
            'this_month_count': this_month_count,
        })
        
        return super().changelist_view(request, extra_context=extra_context)


    # 权限控制方法
    def has_add_permission(self, request):
        """所有销售都可以新增仪器调研"""
        return request.user.is_superuser or request.user.username in ALLOWED_SALESMEN
    
    def has_delete_permission(self, request, obj=None):
        """销售只能删除自己负责医院的仪器"""
        if request.user.is_superuser:
            return True
            
        if not request.user.username in ALLOWED_SALESMEN:
            return False
            
        if obj is None:
            return True
            
        # 检查是否是自己负责的医院的仪器
        return (obj.hospital_survey and 
                obj.hospital_survey.qitian_manager == request.user)
    
    def has_change_permission(self, request, obj=None):
        """销售只能修改自己负责医院的仪器"""
        if request.user.is_superuser:
            return True
            
        if not request.user.username in ALLOWED_SALESMEN:
            return False
            
        if obj is None:
            return True
            
        # 检查是否是自己负责的医院的仪器
        return (obj.hospital_survey and 
                obj.hospital_survey.qitian_manager == request.user)

    #excel
    def _merge_cells(self, ws, merge_ranges):
        """执行单元格合并并设置样式，保留客情颜色"""
        # 客情颜色映射
        familiarity_color_map = {
            'red': 'FFFFE6E6',
            'yellow': 'FFFFFFCC', 
            'green': 'FFE6FFE6',
            'blue': 'FFE6F2FF'
        }
        familiarity_font_color_map = {
            'red': 'FFF40909',
            'yellow': 'FFCC9900',
            'green': 'FF077E07',
            'blue': 'FF3056D2'
        }
        
        for merge_range in merge_ranges:
            start_row = merge_range['start_row']
            end_row = merge_range['end_row']
            start_col = merge_range['start_col']
            end_col = merge_range['end_col']
            merge_type = merge_range['type']
            
            # 只有当需要合并的行数大于1时才执行合并
            if end_row > start_row:
                for col in range(start_col, end_col + 1):
                    # **在合并前先保存客情颜色信息**
                    familiarity_info = None
                    if col == 8 or col == 11:  # 主任客情(第8列)或组长客情(第11列)
                        first_cell = ws.cell(row=start_row, column=col)
                        if first_cell.value in ['不认识', '有商机在跟进', '有明确代理商', '成单']:
                            # 根据值确定颜色key
                            value_to_key = {
                                '不认识': 'red',
                                '有商机在跟进': 'yellow', 
                                '有明确代理商': 'green',
                                '成单': 'blue'
                            }
                            familiarity_info = value_to_key.get(first_cell.value)
                    
                    # 合并单元格
                    ws.merge_cells(
                        start_row=start_row, 
                        start_column=col,
                        end_row=end_row, 
                        end_column=col
                    )
                    
                    # 设置合并单元格的样式
                    merged_cell = ws.cell(row=start_row, column=col)
                    merged_cell.alignment = Alignment(
                        horizontal="center", 
                        vertical="center",
                        wrap_text=True
                    )
                    
                    # **优先设置客情颜色，如果是客情列的话**
                    if familiarity_info and (col == 8 or col == 11):
                        bg_color = familiarity_color_map.get(familiarity_info)
                        font_color = familiarity_font_color_map.get(familiarity_info)
                        
                        if bg_color and font_color:
                            merged_cell.fill = PatternFill(
                                start_color=bg_color, 
                                end_color=bg_color, 
                                fill_type="solid"
                            )
                            merged_cell.font = Font(color=font_color, bold=True)
                    else:
                        # 非客情列才设置默认的合并背景色
                        if merge_type == 'category':
                            merged_cell.fill = PatternFill(
                                start_color="F0F8E8", 
                                end_color="F0F8E8", 
                                fill_type="solid"
                            )
                        elif merge_type == 'instrument':
                            merged_cell.fill = PatternFill(
                                start_color="FFF8E8", 
                                end_color="FFF8E8", 
                                fill_type="solid"
                            )
                        # 医院级别不设置背景色，保持默认

    def _reapply_familiarity_colors(self, ws, data_by_hospital, familiarity_map):
        """重新应用客情颜色（在合并单元格之后）- 增强版本"""
        current_row = 2
        
        # 客情颜色映射
        familiarity_color_map = {
            'red': 'FFFFE6E6',
            'yellow': 'FFFFFFCC',
            'green': 'FFE6FFE6',
            'blue': 'FFE6F2FF'
        }
        familiarity_font_color_map = {
            'red': 'FFF40909',
            'yellow': 'FFCC9900',
            'green': 'FF077E07',
            'blue': 'FF3056D2'
        }
        
        def paint_familiarity_cell(r, c, familiarity_key):
            """为客情单元格设置颜色，处理合并单元格情况"""
            if not familiarity_key:
                return
                
            bg = familiarity_color_map.get(familiarity_key)
            fg = familiarity_font_color_map.get(familiarity_key)
            if not (bg and fg):
                return
                
            target_cell = ws.cell(row=r, column=c)
            
            # 设置背景色
            target_cell.fill = PatternFill(
                start_color=bg.upper(), 
                end_color=bg.upper(), 
                fill_type="solid"
            )
            
            # 设置字体颜色和粗体
            target_cell.font = Font(color=fg.upper(), bold=True)
            
            # 保持居中对齐
            target_cell.alignment = Alignment(
                horizontal="center", 
                vertical="center",
                wrap_text=True
            )
        
        for hospital_survey, categories_data in data_by_hospital.items():
            hospital_start_row = current_row
            
            if not categories_data:
                # 医院没有仪器调研的情况
                # 主任客情（第8列）
                if hospital_survey.director_familiarity:
                    paint_familiarity_cell(current_row, 8, hospital_survey.director_familiarity)
                
                # 组长客情（第11列）
                if hospital_survey.leader_familiarity:
                    paint_familiarity_cell(current_row, 11, hospital_survey.leader_familiarity)
                
                current_row += 1
            else:
                # 有仪器调研的情况
                total_rows_in_hospital = sum(
                    sum(len(projects) if projects else 1 for projects in instruments_data.values())
                    for instruments_data in categories_data.values()
                )
                
                # 为医院的所有行设置客情颜色（因为医院信息会被合并）
                for row_offset in range(total_rows_in_hospital):
                    target_row = hospital_start_row + row_offset
                    
                    # 主任客情（第8列）
                    if hospital_survey.director_familiarity:
                        paint_familiarity_cell(target_row, 8, hospital_survey.director_familiarity)
                    
                    # 组长客情（第11列） 
                    if hospital_survey.leader_familiarity:
                        paint_familiarity_cell(target_row, 11, hospital_survey.leader_familiarity)
                
                current_row += total_rows_in_hospital

                
    def download_comprehensive_excel(self, request, queryset):
        """导出综合仪器调研数据到Excel（带合并单元格）- 显示所有医院调研"""
        
        # 创建工作簿和工作表
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "综合仪器调研数据"
        
        # 定义完整表头
        headers = [
            # 第一部分：医院调研信息（这些字段需要按医院合并）
            '医院', '区域', '级别', '其田负责人', '迈瑞负责人',
            '主任姓名', '主任对接人', '主任客情',
            '组长姓名', '组长对接人', '组长客情',
            '操作老师姓名', '操作老师对接人',
            '销售模式', '分销渠道',
            '血常规标本量', 'CRP标本量', 'SAA标本量', '血沉标本量',
            '糖化标本量', '尿液标本量',
            '血球台数', '糖化台数', '尿液台数',
            '血球品牌-型号-台数-装机年份-标本量',
            '糖化品牌-型号-台数-装机年份-标本量',
            '尿液品牌-型号-台数-装机年份-标本量',
            
            # 第二部分：仪器分类信息
            '仪器分类',
            
            # 第三部分：仪器调研信息
            '是否我司仪器', '我司销售渠道',
            '品牌', '型号', '台数', '装机年份', '安装地点',
            '标本量总和', '血球项目-标本量-竞品关系点-经销商',
            '竞品关系点', '经销商名称',
            
            # 第四部分：血球项目详情
            '血球项目类型', '血球项目标本量', '血球项目竞品关系点', '血球项目经销商',
            
            # 最后：时间信息
            '最后修改人', '更新时间'
        ]
        
        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 写入表头
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 客情显示映射
        familiarity_map = {
            'red': '不认识',
            'yellow': '有商机在跟进', 
            'green': '有明确代理商',
            'blue': '成单'
        }
        
        # **关键修改：基于所有医院调研而不是仪器调研**
        # 获取所有医院调研记录（应用相同的权限和筛选条件）
        # hospital_surveys_queryset = MindrayHospitalSurvey.objects.filter(is_active=True)
        hospital_surveys_queryset = MindrayHospitalSurvey.objects.filter(is_active=True).order_by('id')

      
        # 如果有特定的医院筛选条件，也应用这些条件
        # 从原始queryset中提取医院筛选条件
        if queryset.exists():
            hospital_ids_from_instruments = queryset.values_list('hospital_survey_id', flat=True).distinct()
            # 可以选择是否应用这个筛选，这里我注释掉，显示所有医院
            # hospital_surveys_queryset = hospital_surveys_queryset.filter(id__in=hospital_ids_from_instruments)
        
        # 按医院分组数据（新的方法）
        data_by_hospital = self._group_all_hospital_data(hospital_surveys_queryset)
        
        current_row = 2
        merge_ranges = []
        
        for hospital_survey, categories_data in data_by_hospital.items():
            hospital_start_row = current_row
            
            if not categories_data:
                # 如果医院没有任何仪器调研，仍然显示医院信息
                self._write_data_row(ws, current_row, hospital_survey, None, None, familiarity_map, thin_border)
                current_row += 1
            else:
                # 有仪器调研的情况，按原来的逻辑处理
                for category, instruments_data in categories_data.items():
                    category_start_row = current_row
                    
                    for instrument, projects_data in instruments_data.items():
                        instrument_start_row = current_row
                        
                        if projects_data:
                            for project in projects_data:
                                self._write_data_row(ws, current_row, hospital_survey, instrument, project, familiarity_map, thin_border)
                                current_row += 1
                        else:
                            self._write_data_row(ws, current_row, hospital_survey, instrument, None, familiarity_map, thin_border)
                            current_row += 1
                        
                        # 仪器信息合并范围
                        if len(projects_data) > 1:
                            instrument_end_row = current_row - 1
                            merge_ranges.append({
                                'start_row': instrument_start_row,
                                'end_row': instrument_end_row,
                                'start_col': 29,
                                'end_col': 39,
                                'type': 'instrument'
                            })
                    
                    # 仪器分类合并范围
                    total_rows_in_category = sum(len(projects) if projects else 1 for projects in instruments_data.values())
                    if total_rows_in_category > 1:
                        category_end_row = current_row - 1
                        merge_ranges.append({
                            'start_row': category_start_row,
                            'end_row': category_end_row,
                            'start_col': 28,
                            'end_col': 28,
                            'type': 'category'
                        })
            
            # 医院信息合并范围
            total_rows_in_hospital = 1 if not categories_data else sum(
                sum(len(projects) if projects else 1 for projects in instruments_data.values()) 
                for instruments_data in categories_data.values()
            )
            
            if total_rows_in_hospital > 1:
                hospital_end_row = current_row - 1
                merge_ranges.append({
                    'start_row': hospital_start_row,
                    'end_row': hospital_end_row,
                    'start_col': 1,
                    'end_col': 27,
                    'type': 'hospital'
                })
        
        # 执行单元格合并
        self._merge_cells(ws, merge_ranges)
        
        # 重新应用客情颜色
        self._reapply_familiarity_colors(ws, data_by_hospital, familiarity_map)
        
        # 自动调整列宽
        self._auto_adjust_column_width(ws)
        
        # 冻结首行
        ws.freeze_panes = 'A2'
        
        # 创建HTTP响应
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        current_time = timezone.now().strftime('%Y%m%d_%H%M%S')
        filename = f'综合调研数据_{current_time}.xlsx'
        encoded_filename = quote(filename.encode('utf-8'))
        response['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{encoded_filename}'
        
        wb.save(response)
        
        return response

    download_comprehensive_excel.short_description = "导出汇总"
    download_comprehensive_excel.type = 'success'
    download_comprehensive_excel.style = 'color:white;'

    # def _group_all_hospital_data(self, hospital_surveys_queryset):
    #     """按医院->分类->仪器->项目的层级结构组织数据（包含没有仪器的医院）"""
    #     from collections import defaultdict
        
    #     grouped_data = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
        
    #     for hospital_survey in hospital_surveys_queryset.select_related(
    #         'hospital',
    #         'qitian_manager',
    #         'director_contact',
    #         'leader_contact',
    #         'operator_contact',
    #         'created_by'
    #     ):
    #         # 确保计算最新统计
    #         hospital_survey.calculate_all_statistics()
            
    #         # 获取该医院的所有仪器调研
    #         instruments = MindrayInstrumentSurvey.objects.filter(
    #             hospital_survey=hospital_survey,
    #             is_active=True
    #         ).select_related(
    #             'category',
    #             'brand',
    #             'competitionrelation',
    #             'last_modified_by'
    #         )
            
    #         if not instruments.exists():
    #             # 如果医院没有仪器调研，仍然要显示医院信息
    #             grouped_data[hospital_survey] = {}
    #         else:
    #             # 按仪器分类组织数据
    #             for instrument in instruments:
    #                 category_name = instrument.category.name if instrument.category else '未分类'
                    
    #                 if instrument.category and instrument.category.name == '血球':
    #                     blood_projects = MindrayBloodCellProject.objects.filter(
    #                         instrument_survey=instrument,
    #                         is_active=True
    #                     ).select_related('competitionrelation')
                        
    #                     if blood_projects.exists():
    #                         grouped_data[hospital_survey][category_name][instrument].extend(blood_projects)
    #                     else:
    #                         grouped_data[hospital_survey][category_name][instrument].append(None)
    #                 else:
    #                     grouped_data[hospital_survey][category_name][instrument].append(None)
        
    #     return grouped_data
   
    def _group_all_hospital_data(self, hospital_surveys_queryset):
        """按医院->分类->仪器->项目的层级结构组织数据（包含没有仪器的医院）"""
        from collections import defaultdict
        
        grouped_data = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
        
        for hospital_survey in hospital_surveys_queryset.select_related(
            'hospital',
            'qitian_manager',
            'director_contact',
            'leader_contact',
            'operator_contact',
            'created_by'
        ):
            # 确保计算最新统计
            hospital_survey.calculate_all_statistics()
            
            # **新增：获取该医院的所有仪器调研，应用排序**
            instruments = MindrayInstrumentSurvey.objects.filter(
                hospital_survey=hospital_survey,
                is_active=True
            ).select_related(
                'category',
                'brand',
                'competitionrelation',
                'last_modified_by'
            ).order_by('category__id', 'brand__id','createtime')  # 按分类和品牌排序
            
            if not instruments.exists():
                # 如果医院没有仪器调研，仍然要显示医院信息
                grouped_data[hospital_survey] = {}
            else:
                # 按仪器分类组织数据
                for instrument in instruments:
                    category_name = instrument.category.name if instrument.category else '未分类'
                    
                    if instrument.category and instrument.category.name == '血球':
                        # **新增：血球项目也应用排序（可选）**
                        blood_projects = MindrayBloodCellProject.objects.filter(
                            instrument_survey=instrument,
                            is_active=True
                        ).select_related('competitionrelation').order_by('project_type', 'id')  # 按项目类型排序
                        
                        if blood_projects.exists():
                            grouped_data[hospital_survey][category_name][instrument].extend(blood_projects)
                        else:
                            grouped_data[hospital_survey][category_name][instrument].append(None)
                    else:
                        grouped_data[hospital_survey][category_name][instrument].append(None)
        
        return grouped_data


    def _write_data_row(self, ws, row_num, hospital_survey, instrument, project, familiarity_map, thin_border):
        """写入一行数据（修改以处理空仪器情况）"""
        
        # 计算标本量总和
        total_sample_volume = 0
        if instrument:
            if instrument.category and instrument.category.name == '血球':
                from django.db.models import Sum
                total_sample_volume = MindrayBloodCellProject.objects.filter(
                    instrument_survey=instrument,
                    is_active=True
                ).aggregate(total=Sum('sample_volume'))['total'] or 0
            else:
                total_sample_volume = instrument.sample_volume or 0
        
        # 统一竞品关系点显示
        competition_relation_display = ''
        if instrument:
            if instrument.category and instrument.category.name == '血球':
                competition_relation_display = instrument.blood_competition_relations or '-'
            else:
                competition_relation_display = instrument.competitionrelation.competitionrelation if instrument.competitionrelation else '-'
        
        # 统一经销商显示
        dealer_name_display = ''
        if instrument:
            if instrument.category and instrument.category.name == '血球':
                dealer_name_display = instrument.blood_dealer_names or '-'
            else:
                dealer_name_display = instrument.dealer_name or '-'
        
        # 构建完整行数据
        row_data = [
            # 医院调研信息（列1-27）
            hospital_survey.hospital.hospitalname if hospital_survey and hospital_survey.hospital else '',
            hospital_survey.hospital.district if hospital_survey and hospital_survey.hospital else '',
            hospital_survey.hospital.hospitalclass if hospital_survey and hospital_survey.hospital else '',
            hospital_survey.qitian_manager.chinesename if hospital_survey and hospital_survey.qitian_manager else '',
            hospital_survey.mindray_manager if hospital_survey else '',
            hospital_survey.director_name if hospital_survey else '',
            hospital_survey.director_contact.chinesename if hospital_survey and hospital_survey.director_contact else '',
            familiarity_map.get(hospital_survey.director_familiarity, hospital_survey.director_familiarity or '') if hospital_survey else '',
            hospital_survey.leader_name if hospital_survey else '',
            hospital_survey.leader_contact.chinesename if hospital_survey and hospital_survey.leader_contact else '',
            familiarity_map.get(hospital_survey.leader_familiarity, hospital_survey.leader_familiarity or '') if hospital_survey else '',
            hospital_survey.operator_name if hospital_survey else '',
            hospital_survey.operator_contact.chinesename if hospital_survey and hospital_survey.operator_contact else '',
            dict(MindrayHospitalSurvey.SALES_MODE_CHOICES).get(hospital_survey.sales_mode, hospital_survey.sales_mode or '') if hospital_survey else '',
            hospital_survey.distribution_channel if hospital_survey else '',
            hospital_survey.routine_total_volume if hospital_survey else 0,
            hospital_survey.crp_total_volume if hospital_survey else 0,
            hospital_survey.saa_total_volume if hospital_survey else 0,
            hospital_survey.esr_total_volume if hospital_survey else 0,
            hospital_survey.glycation_total_volume if hospital_survey else 0,
            hospital_survey.urine_total_volume if hospital_survey else 0,
            hospital_survey.blood_cell_total_count if hospital_survey else 0,
            hospital_survey.glycation_total_count if hospital_survey else 0,
            hospital_survey.urine_total_count if hospital_survey else 0,
            hospital_survey.blood_cell_summary if hospital_survey else '',
            hospital_survey.glycation_summary if hospital_survey else '',
            hospital_survey.urine_summary if hospital_survey else '',
            
            # 仪器分类信息（列28）
            instrument.category.name if instrument and instrument.category else '',
            
            # 仪器调研信息（列29-39）
            '是' if instrument and instrument.is_our_instrument else ('否' if instrument else ''),
            dict(MindrayInstrumentSurvey.SALES_CHANNEL_CHOICES).get(instrument.our_sales_channel, instrument.our_sales_channel or '') if instrument and instrument.our_sales_channel else '',
            instrument.brand.brand if instrument and instrument.brand else '',
            instrument.model if instrument else '',
            instrument.quantity if instrument else '',
            instrument.installation_year if instrument else '',
            dict(MindrayInstrumentSurvey.installation_location_CHOICES).get(instrument.installation_location, instrument.installation_location or '') if instrument and instrument.installation_location else '',
            total_sample_volume,
            instrument.blood_project_details or '' if instrument and instrument.category and instrument.category.name == '血球' else '',
            competition_relation_display,
            dealer_name_display,
            
            # 血球项目详情（列40-43）
            project.get_project_type_display() if project else '',
            project.sample_volume or 0 if project else '',
            project.competitionrelation.competitionrelation if project and project.competitionrelation else ('未知' if project else ''),
            project.dealer_name or '未知' if project and project.dealer_name else ('未知' if project else ''),
            
            # 时间信息（列44-45）
            hospital_survey.created_by.chinesename if hospital_survey and hospital_survey.created_by else '',
            self._format_datetime(hospital_survey.updatetime) if hospital_survey else ''
        ]
        
        # 写入数据并设置边框
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    #简单版本
    def download_simple_excel(self, request, queryset):
        """导出当前列表显示的数据到Excel（简单版本）"""
        # 如果没有选中任何项，则导出当前页面的所有数据
        if not queryset.exists():
            queryset = self.get_queryset(request)
            # 应用搜索和过滤条件
            cl = self.get_changelist_instance(request)
            queryset = cl.get_queryset(request)

        queryset = queryset.order_by('hospital_survey__id', 'category__id', 'brand__id','createtime')

        # 创建工作簿和工作表
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "仪器调研数据"
        
        # 定义表头（对应list_display的字段）
        headers = [
            '医院调研',
            '区域', 
            '级别',
            '医院负责人',
            '仪器分类',
            '是否我司仪器',
            '我司业务销售渠道',
            '品牌',
            '型号',
            '台数',
            '装机年份',
            '仪器安装地',
            '标本量总和',
            '血球项目-标本量-竞品关系点-经销商',
            '竞品关系点',
            '经销商名称',
            '最后修改人',
            '更新时间'
        ]
        
        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 写入表头
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 写入数据
        for row_num, obj in enumerate(queryset, 2):
            # 确保在导出前计算最新的统计数据
            if obj.is_blood_category:
                obj.calculate_all_blood_summaries()
            
            # 构建行数据（按照list_display的顺序）
            row_data = [
                str(obj.hospital_survey) if obj.hospital_survey else '',
                self.get_hospital_district(obj),
                self.get_hospital_class(obj),
                self.get_hospital_manager(obj),
                str(obj.category) if obj.category else '',
                '是' if obj.is_our_instrument else '否',
                obj.get_our_sales_channel_display() or '',
                str(obj.brand) if obj.brand else '',
                obj.model or '',
                obj.quantity or 0,
                obj.installation_year or '',
                obj.get_installation_location_display() or '', 
                self.get_total_sample_volume(obj),
                self.get_blood_project_details_display(obj),
                self.get_competition_relation_display(obj),
                self.get_dealer_name_display(obj),
                self.get_last_modified_by(obj),
                self._format_datetime(obj.updatetime)
            ]
            
            # 写入数据并设置边框
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # 自动调整列宽
        self._auto_adjust_column_width(ws)
        
        # 冻结首行
        ws.freeze_panes = 'A2'
        
        # 创建HTTP响应
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # 设置文件名
        current_time = timezone.now().strftime('%Y%m%d_%H%M%S')
        filename = f'仪器调研列表数据_{current_time}.xlsx'
        encoded_filename = quote(filename.encode('utf-8'))
        response['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{encoded_filename}'
        
        # 保存工作簿到响应
        wb.save(response)
        
        return response

    download_simple_excel.short_description = "导出列表"
    download_simple_excel.type = 'warning'
    download_simple_excel.style = 'color:white;'



    class Media:
        css = {
            'all': ('admin/css/freeze_firsttwo_column.css',)
        }
        

@admin.register(MindrayBloodCellProject)
class MindrayBloodCellProjectAdmin(GlobalAdmin):
    list_display = [
        'instrument_survey', 'project_type', 'sample_volume', 
        'competitionrelation', 'dealer_name', 'updatetime'
    ]
    list_filter = ['project_type', 'competitionrelation']
    search_fields = [
        'instrument_survey__hospital_survey__hospital__hospitalname',
        'dealer_name'
    ]
    autocomplete_fields = ['instrument_survey', 'competitionrelation']
    
   
    def delete_model(self, request, obj):
        """删除单个血球项目后触发计算，并同步更新hospital_survey"""
        # 在删除前保存相关引用
        instrument_survey = obj.instrument_survey
        hospital_survey = None
        if instrument_survey and instrument_survey.hospital_survey:
            hospital_survey = instrument_survey.hospital_survey
        
        # 执行删除操作
        super().delete_model(request, obj)
        
        # 重新计算仪器汇总信息
        if instrument_survey:
            instrument_survey.calculate_all_blood_summaries()
            instrument_survey.save(update_fields=[
                'sample_volume', 'blood_project_types', 
                'blood_project_details',               # 新增字段
                'blood_competition_relations', 'blood_dealer_names', 'updatetime'
            ])
        
        # 重新计算全部统计
        if hospital_survey:
            from django.utils import timezone
            hospital_survey.calculate_all_statistics()
            hospital_survey.updatetime = timezone.now()
            hospital_survey.save(update_fields=[
                'crp_total_volume', 'saa_total_volume', 
                'esr_total_volume', 'routine_total_volume',
                'glycation_total_volume', 'urine_total_volume',  # 新增
                'blood_cell_total_count', 'glycation_total_count', 'urine_total_count',
                'blood_cell_summary', 'glycation_summary', 'urine_summary',
                'updatetime'
            ])

    def delete_queryset(self, request, queryset):
        """批量删除血球项目时同步更新hospital_survey的统计信息"""
        # 在删除前收集所有相关的instrument_survey和hospital_survey
        instrument_surveys = set()
        hospital_surveys = set()
        for obj in queryset:
            if obj.instrument_survey:
                instrument_surveys.add(obj.instrument_survey)
                if obj.instrument_survey.hospital_survey:
                    hospital_surveys.add(obj.instrument_survey.hospital_survey)
        
        # 执行批量删除操作
        super().delete_queryset(request, queryset)
        
        # 重新计算所有相关仪器的汇总信息
        for instrument_survey in instrument_surveys:
            instrument_survey.calculate_all_blood_summaries()
            instrument_survey.save(update_fields=[
                'sample_volume', 'blood_project_types', 
                'blood_project_details',               # 新增字段
                'blood_competition_relations', 'blood_dealer_names', 'updatetime'
            ])
        
        # 删除后重新计算并更新所有相关hospital_survey的统计
        from django.utils import timezone
        for hospital_survey in hospital_surveys:
            hospital_survey.calculate_all_statistics()
            hospital_survey.updatetime = timezone.now()
            hospital_survey.save(update_fields=[
                'crp_total_volume', 'saa_total_volume', 
                'esr_total_volume', 'routine_total_volume',
                'glycation_total_volume', 'urine_total_volume',  # 新增
                'blood_cell_total_count', 'glycation_total_count', 'urine_total_count',
                'blood_cell_summary', 'glycation_summary', 'urine_summary',
                'updatetime'
            ])

    def save_model(self, request, obj, form, change):
        """保存单个项目后触发计算，并更新仪器标本量"""
        super().save_model(request, obj, form, change)
        
        # 更新仪器的汇总信息
        if obj.instrument_survey:
            obj.instrument_survey.calculate_all_blood_summaries()
            obj.instrument_survey.save(update_fields=[
                'sample_volume', 'blood_project_types', 
                'blood_project_details',               # 新增字段
                'blood_competition_relations', 'blood_dealer_names', 'updatetime'
            ])
        
        # 触发医院调研的全部统计计算
        if obj.instrument_survey and obj.instrument_survey.hospital_survey:
            from django.utils import timezone
            hospital_survey = obj.instrument_survey.hospital_survey
            hospital_survey.calculate_all_statistics()
            hospital_survey.updatetime = timezone.now()
            hospital_survey.save(update_fields=[
                'crp_total_volume', 'saa_total_volume', 
                'esr_total_volume', 'routine_total_volume',
                'glycation_total_volume', 'urine_total_volume',  # 新增
                'blood_cell_total_count', 'glycation_total_count', 'urine_total_count',
                'blood_cell_summary', 'glycation_summary', 'urine_summary',
                'updatetime'
            ])
 
    # 权限控制方法
    def has_add_permission(self, request):
        """所有销售都可以新增血球项目"""
        return request.user.is_superuser or request.user.username in ALLOWED_SALESMEN
    
    def has_delete_permission(self, request, obj=None):
        """销售只能删除自己负责医院的血球项目"""
        if request.user.is_superuser:
            return True
            
        if not request.user.username in ALLOWED_SALESMEN:
            return False
            
        if obj is None:
            return True
            
        # 检查是否是自己负责的医院的血球项目
        return (obj.instrument_survey and 
                obj.instrument_survey.hospital_survey and
                obj.instrument_survey.hospital_survey.qitian_manager == request.user)
    
    def has_change_permission(self, request, obj=None):
        """销售只能修改自己负责医院的血球项目"""
        if request.user.is_superuser:
            return True
            
        if not request.user.username in ALLOWED_SALESMEN:
            return False
            
        if obj is None:
            return True
            
        # 检查是否是自己负责的医院的血球项目
        return (obj.instrument_survey and 
                obj.instrument_survey.hospital_survey and
                obj.instrument_survey.hospital_survey.qitian_manager == request.user)

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        """根据权限限制外键选择范围"""
        if db_field.name == "instrument_survey":
            kwargs["queryset"] = MindrayInstrumentSurvey.objects.filter(is_active=True)
            
            # 销售只能选择自己负责医院的仪器
            if not request.user.is_superuser and request.user.username in ALLOWED_SALESMEN:
                kwargs["queryset"] = kwargs["queryset"].filter(
                    hospital_survey__qitian_manager=request.user
                )
        
        elif db_field.name == "competitionrelation":
            kwargs["queryset"] = CompetitionRelation.objects.filter(is_active=True)
        
        return super().formfield_for_foreignkey(db_field, request, **kwargs)




#==========================商机==============================

# 销售商机相关的筛选器和Admin配置
class HospitalSurveyManagerFilter(SimpleListFilter):
    """医院调研负责人筛选器（即销售人员筛选）"""
    title = '销售人员'  # 修改标题
    parameter_name = 'hospital_manager'

    def lookups(self, request, model_admin):
        managers = UserInfoMindray.objects.filter(
            Q(username__in=ALLOWED_SALESMEN),
            qitian_manager__isnull=False
        ).distinct()
        return [(manager.id, manager.chinesename) for manager in managers]
    
    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(hospital_survey__qitian_manager__id=self.value())
        return queryset 

 
class OpportunityProjectFilter(SimpleListFilter):
    """商机项目筛选器"""
    title = '商机项目'
    parameter_name = 'opportunity_project'

    def lookups(self, request, model_admin):
        return SalesOpportunity.PROJECT_CHOICES
    
    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(opportunity_project=self.value())
        return queryset


class OpportunityLandingTimeFilter(SimpleListFilter):
    """落地时间筛选器"""
    title = '落地时间'
    parameter_name = 'landing_time'

    def lookups(self, request, model_admin):
        # 获取所有不同的落地时间，按时间倒序
        times = SalesOpportunity.objects.filter(
            is_active=True
        ).values_list('landing_time', flat=True).distinct().order_by('-landing_time')
        
        lookups = []
        for time in times:
            if time:
                lookups.append((time, time))
        return lookups
    
    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(landing_time=self.value())
        return queryset


class OpportunityCreateTimeFilter(SimpleListFilter):
    """创建时间筛选器（本周/上周）"""
    title = '创建时间'
    parameter_name = 'create_time_range'

    def lookups(self, request, model_admin):
        return [
            ('this_week', '本周新增'),
            ('last_week', '上周新增'),
            ('this_month', '本月新增'),
            ('last_month', '上月新增'),
        ]
    
    def queryset(self, request, queryset):
        if self.value():
            from django.utils import timezone
            from datetime import timedelta
            
            today = timezone.now().date()
            
            if self.value() == 'this_week':
                week_start = today - timedelta(days=today.weekday())
                return queryset.filter(createtime__date__gte=week_start)
            
            elif self.value() == 'last_week':
                this_week_start = today - timedelta(days=today.weekday())
                last_week_start = this_week_start - timedelta(days=7)
                last_week_end = this_week_start - timedelta(days=1)
                return queryset.filter(
                    createtime__date__gte=last_week_start,
                    createtime__date__lte=last_week_end
                )
            
            elif self.value() == 'this_month':
                month_start = today.replace(day=1)
                return queryset.filter(createtime__date__gte=month_start)
            
            elif self.value() == 'last_month':
                from calendar import monthrange
                if today.month == 1:
                    last_month = today.replace(year=today.year-1, month=12, day=1)
                else:
                    last_month = today.replace(month=today.month-1, day=1)
                
                # 上个月的最后一天
                days_in_month = monthrange(last_month.year, last_month.month)[1]
                last_month_end = last_month.replace(day=days_in_month)
                
                return queryset.filter(
                    createtime__date__gte=last_month,
                    createtime__date__lte=last_month_end
                )
        
        return queryset


class SalesOpportunityForm(forms.ModelForm):
    class Meta:
        model = SalesOpportunity
        fields = "__all__"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        # 把 help_text 设置为 placeholder
        if "landing_time" in self.fields:
            self.fields["landing_time"].widget.attrs.update({
                "placeholder": self.fields["landing_time"].help_text
            })
            # 可选：去掉外部 help_text，避免重复显示
            self.fields["landing_time"].help_text = ""


@admin.register(SalesOpportunity)
class SalesOpportunityAdmin(GlobalAdmin):
    form = SalesOpportunityForm
    ordering = ['hospital_survey__id','opportunity_project','-createtime']
    autocomplete_fields = ['hospital_survey']  # 移除salesperson
    
    list_display = [
        'get_hospital_name',
        'get_hospital_district',
        'get_hospital_class', 
        'get_hospital_manager',  # 这就是销售人员，不需要单独显示salesperson了
        'opportunity_model_short',
        'get_opportunity_project_colored',      
        'sample_volume',
        'landing_time',
        'opportunity_status_short',
        'get_create_week',
        'createtime',
        'updatetime'
    ]
    
    list_filter = [
        'hospital_survey__hospital__district',
        'hospital_survey__hospital__hospitalclass',
        HospitalSurveyManagerFilter,  # 这个筛选器实际上就是按销售人员筛选
        OpportunityProjectFilter,
        OpportunityLandingTimeFilter,
        OpportunityCreateTimeFilter,
    ]
    
    search_fields = [
        'hospital_survey__hospital__hospitalname',
        'hospital_survey__qitian_manager__chinesename',
        'opportunity_model',
        'opportunity_status',
        'landing_time'
    ]
    
    fieldsets = (
        ('基本信息', {
            'fields': ('hospital_survey',),  # 只显示医院调研选择
            'classes': ('primary-fieldset',)
        }),
        ('商机详情', {
            'fields': (
                ('opportunity_model','opportunity_project'),
                ('sample_volume', 'landing_time'),
                'opportunity_status'
            ),
            'classes': ('secondary-fieldset',)
        }),
      
    )
    
    readonly_fields = ['createtime', 'updatetime', 'get_auto_salesperson_display']
    
    #项目分类变成有颜色的span
    def get_opportunity_project_colored(self, obj):
        """显示带颜色和背景的商机项目"""
        project_display = obj.get_opportunity_project_display()
        
        # 定义样式映射
        style_mapping = {
            '血球': {
                'color': '#FFFFFF',
                'background': '#FF6B6B',
                'border': '1px solid #FF6B6B'
            },
            '糖化': {
                'color': '#FFFFFF', 
                'background': '#F5A512',
                'border': '1px solid F5A512'
            },
            '尿液': {
                'color': '#FFFFFF',
                'background': '#45B7D1', 
                'border': '1px solid #45B7D1'
            }
        }
        
        # 获取对应样式，默认样式
        styles = style_mapping.get(project_display, {
            'color': '#333333',
            'background': '#F8F9FA',
            'border': '1px solid #DEE2E6'
        })
        
        style_str = f"color: {styles['color']}; background-color: {styles['background']}; border: {styles['border']}; padding: 3px 8px; border-radius: 4px; font-weight: bold; font-size: 12px;"
        
        return format_html(
            '<span style="{}">{}</span>',
            style_str,
            project_display
        )

    get_opportunity_project_colored.short_description = '商机项目'
    get_opportunity_project_colored.admin_order_field = 'opportunity_project'

    # 显示方法
    def get_hospital_name(self, obj):
        if obj.hospital_survey and obj.hospital_survey.hospital:
            return obj.hospital_survey.hospital.hospitalname
        return '-'
    get_hospital_name.short_description = '医院名称'
    get_hospital_name.admin_order_field = 'hospital_survey__hospital__hospitalname'
    
    def get_hospital_district(self, obj):
        if obj.hospital_survey and obj.hospital_survey.hospital:
            return obj.hospital_survey.hospital.district
        return '-'
    get_hospital_district.short_description = '区域'
    get_hospital_district.admin_order_field = 'hospital_survey__hospital__district'
    
    def get_hospital_class(self, obj):
        if obj.hospital_survey and obj.hospital_survey.hospital:
            return obj.hospital_survey.hospital.hospitalclass
        return '-'
    get_hospital_class.short_description = '级别'
    get_hospital_class.admin_order_field = 'hospital_survey__hospital__hospitalclass'
    
    def get_hospital_manager(self, obj):
        """显示销售人员（自动从医院调研负责人获取）"""
        if obj.hospital_survey and obj.hospital_survey.qitian_manager:
            return obj.hospital_survey.qitian_manager.chinesename
        return '-'
    get_hospital_manager.short_description = '销售人员'  # 改名为销售人员
    get_hospital_manager.admin_order_field = 'hospital_survey__qitian_manager__chinesename'
    
    def get_auto_salesperson_display(self, obj):
        """在表单中显示自动设置的销售人员"""
        if obj.hospital_survey and obj.hospital_survey.qitian_manager:
            return f"{obj.hospital_survey.qitian_manager.chinesename} (自动设置)"
        return "将根据选择的医院调研自动设置"
    get_auto_salesperson_display.short_description = '销售人员'
    
    def opportunity_model_short(self, obj):
        """显示商机型号的简短版本"""
        if len(obj.opportunity_model) > 20:
            return obj.opportunity_model[:20] + '...'
        return obj.opportunity_model
    opportunity_model_short.short_description = '商机型号'
    opportunity_model_short.admin_order_field = 'opportunity_model'
    
    def opportunity_status_short(self, obj):
        """显示商机情况的简短版本"""
        if obj.opportunity_status and len(obj.opportunity_status) > 30:
            return obj.opportunity_status[:30] + '...'
        return obj.opportunity_status or '-'
    opportunity_status_short.short_description = '商机情况'
    opportunity_status_short.admin_order_field = 'opportunity_status'
    
    def get_create_week(self, obj):
        """显示创建是第几周"""
        if obj.createtime:
            week_num = obj.createtime.isocalendar()[1]
            return f"第{week_num}周"
        return '-'
    get_create_week.short_description = '创建周次'
    
    # 权限控制
    def get_queryset(self, request):
        """根据用户权限过滤数据"""
        qs = super().get_queryset(request)
        qs = qs.filter(is_active=True)
        
        # # 普通销售只能看到自己负责的医院调研相关的商机
        # if not request.user.is_superuser and request.user.username in ALLOWED_SALESMEN:
        #     qs = qs.filter(hospital_survey__qitian_manager=request.user)
        
        return qs.select_related(
            'hospital_survey__hospital',
            'hospital_survey__qitian_manager', 
            'salesperson'
        )
    

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        """表单外键字段的查询集控制"""
        if db_field.name == "hospital_survey":
            kwargs["queryset"] = MindrayHospitalSurvey.objects.filter(is_active=True)
            
            # 销售只能选择自己负责的医院调研
            if not request.user.is_superuser and request.user.username in ALLOWED_SALESMEN:
                kwargs["queryset"] = kwargs["queryset"].filter(qitian_manager=request.user)
        
        return super().formfield_for_foreignkey(db_field, request, **kwargs)


    def changelist_view(self, request, extra_context=None):
        """自定义列表页面，添加统计信息和图表数据"""
        extra_context = extra_context or {}
        
        queryset = self.get_queryset(request)
        cl = self.get_changelist_instance(request)
        queryset = cl.get_queryset(request)
        
        from django.utils import timezone
        from datetime import timedelta
        import json
        
        today = timezone.now().date()
        week_start = today - timedelta(days=today.weekday())
        
        this_week_count = queryset.filter(createtime__date__gte=week_start).count()
        
        last_week_start = week_start - timedelta(days=7)
        last_week_end = week_start - timedelta(days=1)
        last_week_count = queryset.filter(
            createtime__date__gte=last_week_start,
            createtime__date__lte=last_week_end
        ).count()
        
        total_count = queryset.count()
 
        project_stats = {
            '血球': queryset.filter(opportunity_project='BLOOD_CELL').count(),
            '糖化': queryset.filter(opportunity_project='GLYCATION').count(),
            '尿液': queryset.filter(opportunity_project='URINE').count(),
        }
        # 准备图表数据 - 只包含有数据的项目
        project_chart_data = []
        project_chart_labels = []
        project_chart_colors = []
        
       
        color_mapping = {
            '血球': '#FF6B6B',
            '糖化': '#4ECDC4', 
            '尿液': '#45B7D1',
        }
                
        for project_name, count in project_stats.items():
            if count > 0:
                project_chart_data.append(count)
                project_chart_labels.append(project_name)
                project_chart_colors.append(color_mapping[project_name])
        
        # 2. Top10医院数据
        top_hospitals_data = queryset.values(
            'hospital_survey__hospital__hospitalname',
            'hospital_survey__hospital__district'
        ).annotate(
            count=Count('id')
        ).order_by('-count')[:10]
        
        top_hospitals_list = []
        for i, item in enumerate(top_hospitals_data, 1):
            hospital_name = item['hospital_survey__hospital__hospitalname'] or '未知医院'
            district = item['hospital_survey__hospital__district'] or '未知区域'
            count = item['count']
            top_hospitals_list.append({
                'rank': i,
                'hospital_name': hospital_name,
                'district': district,
                'count': count
            })
        
        # 3. 销售排行数据
        salesperson_stats = queryset.values(
            'hospital_survey__qitian_manager__chinesename'
        ).annotate(
            count=Count('id')
        ).order_by('-count')[:10]
        
        salesperson_names = []
        salesperson_counts = []
        
        for item in salesperson_stats:
            name = item['hospital_survey__qitian_manager__chinesename'] or '未分配'
            count = item['count']
            salesperson_names.append(name)
            salesperson_counts.append(count)
        
        # 按区域统计
        district_stats = queryset.values(
            'hospital_survey__hospital__district'
        ).annotate(
            count=Count('id'),
            total_volume=Sum('sample_volume')
        ).order_by('-count')
        
        extra_context.update({
            'total_count': total_count,
            'this_week_count': this_week_count,
            'last_week_count': last_week_count,
            'district_stats': district_stats,
            'is_salesperson': not request.user.is_superuser and request.user.username in ALLOWED_SALESMEN,
            
            # 原始数据 - 用于JavaScript
            'project_chart_data': project_chart_data,
            'project_chart_labels': project_chart_labels, 
            'project_chart_colors': project_chart_colors,
            'salesperson_names': salesperson_names,
            'salesperson_counts': salesperson_counts,
            
            # 医院排名
            'top_hospitals_list': top_hospitals_list,
            
            # 调试信息
            'project_stats_debug': project_stats,
        })
        
        return super().changelist_view(request, extra_context=extra_context)


    def save_model(self, request, obj, form, change):
        """只有在实际有改动时才更新时间和触发联动更新"""
        
        # 检查是否有实际改动
        has_changes = False
        if not change:  # 新建记录
            has_changes = True
        elif form.changed_data:  # 修改且有字段变化
            has_changes = True
        
        if not has_changes and change:
            # 没有实际改动，只刷新数据不保存
            obj.refresh_from_db()
            return
        
        # 如果没有改动，使用update_fields避免触发auto_now
        if change and not has_changes:
            obj.refresh_from_db()
            return
        
        super().save_model(request, obj, form, change)
        
        # 只有在有实际改动时才触发医院调研的商机汇总重新计算
        if has_changes and obj.hospital_survey:
            # 重新计算医院调研的商机汇总
            obj.hospital_survey.calculate_all_statistics()
            
            # 手动更新hospital_survey的updatetime
            from django.utils import timezone
            obj.hospital_survey.updatetime = timezone.now()
            
            obj.hospital_survey.save(update_fields=[
                'crp_total_volume', 'saa_total_volume', 
                'esr_total_volume', 'routine_total_volume',
                'glycation_total_volume', 'urine_total_volume',
                'blood_cell_total_count', 'glycation_total_count', 'urine_total_count',
                'blood_cell_summary', 'glycation_summary', 'urine_summary',
                'sales_opportunities_summary',  # 商机汇总字段
                'updatetime'
            ])
    
 
    def delete_model(self, request, obj):
        """删除单个商机时同步更新hospital_survey的商机汇总"""
        # 在删除前先保存hospital_survey引用
        hospital_survey = obj.hospital_survey
        
        # 执行删除操作（逻辑删除，设置is_active=False）
        obj.is_active = False
        obj.save()
        
        # 删除后重新计算并更新hospital_survey的商机汇总
        if hospital_survey:
            from django.utils import timezone
            hospital_survey.calculate_all_statistics()
            hospital_survey.updatetime = timezone.now()
            hospital_survey.save(update_fields=[
                'crp_total_volume', 'saa_total_volume', 
                'esr_total_volume', 'routine_total_volume',
                'glycation_total_volume', 'urine_total_volume',
                'blood_cell_total_count', 'glycation_total_count', 'urine_total_count',
                'blood_cell_summary', 'glycation_summary', 'urine_summary',
                'sales_opportunities_summary',
                'updatetime'
            ])

    def delete_queryset(self, request, queryset):
        """批量删除商机时同步更新hospital_survey的商机汇总"""
        # 在删除前收集所有相关的hospital_survey
        hospital_surveys = set()
        for obj in queryset:
            if obj.hospital_survey:
                hospital_surveys.add(obj.hospital_survey)
        
        # 执行批量删除操作
        queryset.update(is_active=False)
        
        # 删除后重新计算并更新所有相关hospital_survey的商机汇总
        from django.utils import timezone
        for hospital_survey in hospital_surveys:
            hospital_survey.calculate_all_statistics()
            hospital_survey.updatetime = timezone.now()
            hospital_survey.save(update_fields=[
                'crp_total_volume', 'saa_total_volume', 
                'esr_total_volume', 'routine_total_volume',
                'glycation_total_volume', 'urine_total_volume',
                'blood_cell_total_count', 'glycation_total_count', 'urine_total_count',
                'blood_cell_summary', 'glycation_summary', 'urine_summary',
                'sales_opportunities_summary',
                'updatetime'
            ])

 
    actions = ['download_opportunity_excel', 'export_opportunity_summary']
    
    def download_opportunity_excel(self, request, queryset):
        """导出选中的商机数据到Excel"""
        if not queryset.exists():
            queryset = self.get_queryset(request)

        queryset = queryset.order_by('hospital_survey__id', 'opportunity_project', '-createtime')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "销售商机数据"
        
        # 定义表头（移除重复的销售人员列）
        headers = [
            '医院名称', '区域', '级别', '销售人员', 
            '商机型号', '商机项目', '商机标本量', '落地时间', '商机情况',
            '创建时间', '更新时间'
        ]
        
        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 写入表头
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 写入数据
        for row_num, obj in enumerate(queryset, 2):
            create_time_str = ''
            update_time_str = ''
            if obj.createtime:
                local_create_time = timezone.localtime(obj.createtime)
                create_time_str = local_create_time.strftime('%Y-%m-%d %H:%M:%S')
            if obj.updatetime:
                local_update_time = timezone.localtime(obj.updatetime)
                update_time_str = local_update_time.strftime('%Y-%m-%d %H:%M:%S')
            
            row_data = [
                obj.hospital_survey.hospital.hospitalname if obj.hospital_survey and obj.hospital_survey.hospital else '',
                obj.hospital_survey.hospital.district if obj.hospital_survey and obj.hospital_survey.hospital else '',
                obj.hospital_survey.hospital.hospitalclass if obj.hospital_survey and obj.hospital_survey.hospital else '',
                obj.hospital_survey.qitian_manager.chinesename if obj.hospital_survey and obj.hospital_survey.qitian_manager else '',  # 销售人员（自动获取）
                obj.opportunity_model or '',
                obj.get_opportunity_project_display(),
                obj.sample_volume or 0,
                obj.landing_time or '',
                obj.opportunity_status or '',
                create_time_str,
                update_time_str
            ]
            
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        
        # 自动调整列宽等其余代码保持不变...
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        ws.freeze_panes = 'A2'
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        current_time = timezone.now().strftime('%Y%m%d_%H%M%S')
        filename = f'销售商机数据_{current_time}.xlsx'
        encoded_filename = quote(filename.encode('utf-8'))
        response['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{encoded_filename}'
        
        wb.save(response)
        return response
    
    download_opportunity_excel.short_description = "导出商机列表"
    download_opportunity_excel.type = 'warning'
    download_opportunity_excel.style = 'color:white;'
    

    def export_opportunity_summary(self, request, queryset):
        """导出商机汇总（数据透视表格式）"""
        if not queryset.exists():
            queryset = self.get_queryset(request)
        
        # 按照admin的排序规则排序
        queryset = queryset.order_by('hospital_survey__id', 'opportunity_project', '-createtime')
        
        # 获取所有不同的落地时间，按时间排序
        landing_times = sorted(list(set(
            obj.landing_time for obj in queryset 
            if obj.landing_time
        )))
        
        # 构建数据结构：按医院、商机型号、商机项目分组
        summary_data = {}
        
        for obj in queryset:
            if not (obj.hospital_survey and obj.hospital_survey.hospital):
                continue
                
            hospital = obj.hospital_survey.hospital
            hospital_key = (
                hospital.hospitalname,
                hospital.district or '',
                hospital.hospitalclass or '',
                obj.hospital_survey.qitian_manager.chinesename if obj.hospital_survey.qitian_manager else ''
            )
            
            project_key = (
                obj.opportunity_model or '',
                obj.get_opportunity_project_display() or '',
                obj.opportunity_status or '',  # 添加商机情况
                obj.createtime.strftime('%Y-%m-%d') if obj.createtime else ''
            )
            
            full_key = hospital_key + project_key
            
            if full_key not in summary_data:
                summary_data[full_key] = {
                    'hospital_name': hospital_key[0],
                    'district': hospital_key[1], 
                    'hospital_class': hospital_key[2],
                    'salesperson': hospital_key[3],
                    'opportunity_model': project_key[0],
                    'opportunity_project': project_key[1],
                    'opportunity_status': project_key[2],  # 添加商机情况
                    'create_time': project_key[3],  # 索引位置调整
                    'landing_volumes': {},
                    'total_volume': 0,
                    # 添加排序辅助字段
                    'hospital_survey_id': obj.hospital_survey.id if obj.hospital_survey else 0,
                    'opportunity_project_order': obj.opportunity_project,
                    'createtime': obj.createtime
                }
            
            # 按落地时间汇总标本量
            landing_time = obj.landing_time or '未设置'
            volume = obj.sample_volume or 0
            
            if landing_time not in summary_data[full_key]['landing_volumes']:
                summary_data[full_key]['landing_volumes'][landing_time] = 0
            
            summary_data[full_key]['landing_volumes'][landing_time] += volume
            summary_data[full_key]['total_volume'] += volume
        
        # 对汇总数据按照相同的排序规则排序
        sorted_items = sorted(summary_data.items(), key=lambda x: (
            x[1]['hospital_survey_id'],
            x[1]['opportunity_project_order'] or '',
            x[1]['createtime'] or timezone.now()
        ))
        
        # 创建Excel工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "商机汇总"
        
        # 构建表头 - 添加商机情况列
        base_headers = ['医院', '区域', '级别', '销售人员', '商机型号', '商机项目', '商机情况', '创建时间']
        landing_time_headers = [str(lt) for lt in landing_times]
        if '未设置' not in landing_time_headers:
            landing_time_headers.append('未设置')
        
        all_headers = base_headers + landing_time_headers + ['总计']
        
        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 写入表头
        for col_num, header in enumerate(all_headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 按医院分组数据，用于合并单元格 - 使用排序后的数据
        hospital_groups = {}
        for key, data in sorted_items:
            hospital_info = (data['hospital_name'], data['district'], 
                            data['hospital_class'], data['salesperson'])
            if hospital_info not in hospital_groups:
                hospital_groups[hospital_info] = []
            hospital_groups[hospital_info].append((key, data))
        
        # 写入数据
        current_row = 2
        total_by_landing_time = {}  # 用于计算各落地时间的总计
        grand_total = 0  # 总计
        
        # 按排序后的医院顺序处理
        processed_hospitals = set()
        for key, data in sorted_items:
            hospital_info = (data['hospital_name'], data['district'], 
                            data['hospital_class'], data['salesperson'])
            
            if hospital_info in processed_hospitals:
                continue
            
            processed_hospitals.add(hospital_info)
            projects = hospital_groups[hospital_info]
            hospital_start_row = current_row
            
            # 写入该医院的所有项目数据
            for proj_key, proj_data in projects:
                # 基础信息列
                ws.cell(row=current_row, column=1, value=proj_data['hospital_name'])
                ws.cell(row=current_row, column=2, value=proj_data['district']) 
                ws.cell(row=current_row, column=3, value=proj_data['hospital_class'])
                ws.cell(row=current_row, column=4, value=proj_data['salesperson'])
                ws.cell(row=current_row, column=5, value=proj_data['opportunity_model'])
                ws.cell(row=current_row, column=6, value=proj_data['opportunity_project'])
                ws.cell(row=current_row, column=7, value=proj_data['opportunity_status'])  # 添加商机情况
                ws.cell(row=current_row, column=8, value=proj_data['create_time'])  # 列数后移
                
                # 各落地时间的标本量 - 起始列数调整为9
                col_num = 9
                for landing_time in landing_time_headers:
                    volume = proj_data['landing_volumes'].get(landing_time, 0)
                    if volume > 0:
                        ws.cell(row=current_row, column=col_num, value=volume)
                    
                    # 累加到总计中
                    if landing_time not in total_by_landing_time:
                        total_by_landing_time[landing_time] = 0
                    total_by_landing_time[landing_time] += volume
                    
                    col_num += 1
                
                # 该项目总计
                ws.cell(row=current_row, column=len(all_headers), value=proj_data['total_volume'])
                grand_total += proj_data['total_volume']
                
                current_row += 1
            
            # 合并医院相关的单元格（1-4列）
            if len(projects) > 1:
                hospital_end_row = current_row - 1
                for col in range(1, 5):  # 列1-4：医院、区域、级别、销售人员
                    ws.merge_cells(
                        start_row=hospital_start_row, 
                        start_column=col,
                        end_row=hospital_end_row, 
                        end_column=col
                    )
                    # 设置合并单元格的对齐方式
                    merged_cell = ws.cell(row=hospital_start_row, column=col)
                    merged_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 添加总计行
        total_row = current_row + 1
        ws.cell(row=total_row, column=1, value="总计").font = Font(bold=True)
        
        # 各落地时间的总计 - 起始列数调整为9
        col_num = 9
        for landing_time in landing_time_headers:
            total_volume = total_by_landing_time.get(landing_time, 0)
            if total_volume > 0:
                cell = ws.cell(row=total_row, column=col_num, value=total_volume)
                cell.font = Font(bold=True)
            col_num += 1
        
        # 总的总计
        grand_total_cell = ws.cell(row=total_row, column=len(all_headers), value=grand_total)
        grand_total_cell.font = Font(bold=True)
        
        # 设置总计行的背景色
        total_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        for col in range(1, len(all_headers) + 1):
            ws.cell(row=total_row, column=col).fill = total_fill
        
        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max(max_length + 2, 10), 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 冻结首行和前4列
        ws.freeze_panes = 'E2'
        
        # 设置所有数据的边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(1, total_row + 1):
            for col in range(1, len(all_headers) + 1):
                ws.cell(row=row, column=col).border = thin_border
        
        # 创建响应
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        current_time = timezone.now().strftime('%Y%m%d_%H%M%S')
        filename = f'商机汇总_{current_time}.xlsx'
        encoded_filename = quote(filename.encode('utf-8'))
        response['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{encoded_filename}'
        
        wb.save(response)
        return response

    export_opportunity_summary.short_description = "导出商机汇总"
    export_opportunity_summary.type = 'success'
    export_opportunity_summary.style = 'color:white;'

 
   # 权限控制方法
    def has_add_permission(self, request):
        """所有销售都可以新增商机"""
        return request.user.is_superuser or request.user.username in ALLOWED_SALESMEN
    
    def has_delete_permission(self, request, obj=None):
        """销售只能删除自己负责医院的商机"""
        if request.user.is_superuser:
            return True
            
        if not request.user.username in ALLOWED_SALESMEN:
            return False
            
        if obj is None:
            return True
            
        # 检查是否是自己负责的医院的商机
        return (obj.hospital_survey and 
                obj.hospital_survey.qitian_manager == request.user)
    
    def has_change_permission(self, request, obj=None):
        """销售只能修改自己负责医院的商机"""
        if request.user.is_superuser:
            return True
            
        if not request.user.username in ALLOWED_SALESMEN:
            return False
            
        if obj is None:
            return True
            
        # 检查是否是自己负责的医院的商机
        return (obj.hospital_survey and 
                obj.hospital_survey.qitian_manager == request.user)


#================================商机summary================================
 
class MultiSelectMonthFilter(SimpleListFilter):
    """多选月筛选器"""
    title = '月份筛选'
    parameter_name = 'months'

    def lookups(self, request, model_admin):
        from django.utils import timezone
        current_year = timezone.now().year
        months = []
        
        # 当前年份的月份（倒序）
        for month in range(12, 0, -1):
            month_name = f"{current_year}年{month}月"
            months.append((f"{current_year}-{month:02d}", month_name))
        
        # 如果有去年的数据，添加去年的月份
        last_year = current_year - 1
        if SalesOpportunity.objects.filter(createtime__year=last_year).exists():
            for month in range(12, 0, -1):
                month_name = f"{last_year}年{month}月"
                months.append((f"{last_year}-{month:02d}", month_name))
        
        return months

    def queryset(self, request, queryset):
        selected_months = request.GET.getlist(self.parameter_name)
        if selected_months and selected_months != ['']:
            from django.db.models import Q
            
            month_conditions = Q()
            for month_str in selected_months:
                if '-' in month_str and month_str != '':
                    try:
                        year, month = month_str.split('-')
                        month_conditions |= Q(
                            createtime__year=int(year),
                            createtime__month=int(month)
                        )
                    except (ValueError, IndexError):
                        continue
            
            if month_conditions:
                hospital_surveys_with_opportunities = SalesOpportunity.objects.filter(
                    month_conditions, is_active=True
                ).values_list('hospital_survey_id', flat=True).distinct()
                
                # 修改这里：使用 hospital_survey_id 而不是 id
                return queryset.filter(hospital_survey_id__in=hospital_surveys_with_opportunities)
        
        return queryset

    def choices(self, changelist):
        """重写choices方法以支持多选"""
        selected_values = changelist.request.GET.getlist(self.parameter_name)
        
        yield {
            'selected': False,
            'query_string': changelist.get_query_string(remove=[self.parameter_name]),
            'display': '全部月份',
            'reset_link': True
        }
        
        for lookup, title in self.lookup_choices:
            yield {
                'selected': lookup in selected_values,
                'query_string': self._get_query_string(changelist, lookup, selected_values),
                'display': title,
                'lookup': lookup
            }
    
    def _get_query_string(self, changelist, lookup_value, current_selected):
        """生成多选的查询字符串"""
        new_selected = list(current_selected)
        
        if lookup_value in current_selected:
            new_selected.remove(lookup_value)
        else:
            new_selected.append(lookup_value)
        
        if new_selected:
            return changelist.get_query_string({self.parameter_name: new_selected})
        else:
            return changelist.get_query_string(remove=[self.parameter_name])


class MultiSelectYearFilter(SimpleListFilter):
    """多选年筛选器"""
    title = '年份筛选'
    parameter_name = 'years'

    def lookups(self, request, model_admin):
        # 获取所有商机的年份（倒序）
        years = SalesOpportunity.objects.filter(
            is_active=True
        ).dates('createtime', 'year', order='DESC').distinct()
        
        return [(year.year, f"{year.year}年") for year in years]

    def queryset(self, request, queryset):
        selected_years = request.GET.getlist(self.parameter_name)
        if selected_years and selected_years != ['']:
            from django.db.models import Q
            
            year_conditions = Q()
            for year in selected_years:
                if year and year != '':
                    try:
                        year_conditions |= Q(createtime__year=int(year))
                    except ValueError:
                        continue
            
            if year_conditions:
                hospital_surveys_with_opportunities = SalesOpportunity.objects.filter(
                    year_conditions, is_active=True
                ).values_list('hospital_survey_id', flat=True).distinct()
                
                # 修改这里：使用 hospital_survey_id 而不是 id
                return queryset.filter(hospital_survey_id__in=hospital_surveys_with_opportunities)
        
        return queryset

    def choices(self, changelist):
        """重写choices方法以支持多选"""
        selected_values = changelist.request.GET.getlist(self.parameter_name)
        
        yield {
            'selected': False,
            'query_string': changelist.get_query_string(remove=[self.parameter_name]),
            'display': '全部年份',
            'reset_link': True
        }
        
        for lookup, title in self.lookup_choices:
            yield {
                'selected': lookup in selected_values,
                'query_string': self._get_query_string(changelist, lookup, selected_values),
                'display': title,
                'lookup': lookup
            }
    
    def _get_query_string(self, changelist, lookup_value, current_selected):
        """生成多选的查询字符串"""
        new_selected = list(current_selected)
        
        if lookup_value in current_selected:
            new_selected.remove(lookup_value)
        else:
            new_selected.append(lookup_value)
        
        if new_selected:
            return changelist.get_query_string({self.parameter_name: new_selected})
        else:
            return changelist.get_query_string(remove=[self.parameter_name]) 
 

class SummaryHospitalManagerFilter(SimpleListFilter):
    """汇总页面的销售人员筛选器"""
    title = '销售人员'
    parameter_name = 'salesperson_id'

    def lookups(self, request, model_admin):
        managers = UserInfoMindray.objects.filter(
            Q(username__in=ALLOWED_SALESMEN),
            sales_opportunities__isnull=False  # 只显示有商机的销售人员
        ).distinct()
        return [(manager.id, manager.chinesename) for manager in managers]
    
    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(salesperson__id=self.value())
        return queryset


class SummaryDateRangeFilter(SimpleListFilter):
    """汇总页面的日期范围筛选器"""
    title = '商机创建时间'
    parameter_name = 'date_range'
    
    def lookups(self, request, model_admin):
        return [
            ('today', '今天'),
            ('yesterday', '昨天'),
            ('this_week', '本周'),
            ('last_week', '上周'),
            ('this_month', '本月'),
            ('last_month', '上月'),
            ('this_quarter', '本季度'),
            ('this_year', '今年'),
            ('custom', '自定义时间范围'),
        ]

    def queryset(self, request, queryset):
        from django.utils import timezone
        from datetime import timedelta
        import calendar
        
        # 处理自定义日期范围
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        
        if start_date and end_date:
            try:
                from datetime import datetime
                start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
                end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
                
                # 获取在该时间范围内有商机的医院调研
                hospital_surveys_with_opportunities = SalesOpportunity.objects.filter(
                    createtime__date__gte=start_date_obj,
                    createtime__date__lte=end_date_obj,
                    is_active=True
                ).values_list('hospital_survey_id', flat=True).distinct()
                
                return queryset.filter(hospital_survey_id__in=hospital_surveys_with_opportunities)
                
            except ValueError:
                pass
        
        # 处理快捷选项
        value = self.value()
        if not value:
            return queryset
            
        today = timezone.now().date()
        
        if value == 'today':
            start_date_obj = end_date_obj = today
        elif value == 'yesterday':
            yesterday = today - timedelta(days=1)
            start_date_obj = end_date_obj = yesterday
        elif value == 'this_week':
            start_date_obj = today - timedelta(days=today.weekday())
            end_date_obj = start_date_obj + timedelta(days=6)
        elif value == 'last_week':
            start_of_this_week = today - timedelta(days=today.weekday())
            start_date_obj = start_of_this_week - timedelta(days=7)
            end_date_obj = start_of_this_week - timedelta(days=1)
        elif value == 'this_month':
            start_date_obj = today.replace(day=1)
            end_date_obj = today.replace(day=calendar.monthrange(today.year, today.month)[1])
        elif value == 'last_month':
            if today.month == 1:
                last_month = today.replace(year=today.year-1, month=12, day=1)
            else:
                last_month = today.replace(month=today.month-1, day=1)
            start_date_obj = last_month
            end_date_obj = last_month.replace(day=calendar.monthrange(last_month.year, last_month.month)[1])
        elif value == 'this_quarter':
            quarter = (today.month - 1) // 3 + 1
            start_date_obj = today.replace(month=(quarter-1)*3+1, day=1)
            if quarter == 4:
                end_date_obj = today.replace(month=12, day=31)
            else:
                end_date_obj = today.replace(month=quarter*3, day=calendar.monthrange(today.year, quarter*3)[1])
        elif value == 'this_year':
            start_date_obj = today.replace(month=1, day=1)
            end_date_obj = today.replace(month=12, day=31)
        else:
            return queryset
        
        # 获取在该时间范围内有商机的医院调研
        hospital_surveys_with_opportunities = SalesOpportunity.objects.filter(
            createtime__date__gte=start_date_obj,
            createtime__date__lte=end_date_obj,
            is_active=True
        ).values_list('hospital_survey_id', flat=True).distinct()
        
        return queryset.filter(hospital_survey_id__in=hospital_surveys_with_opportunities)

    def choices(self, changelist):
        """重写choices方法以添加自定义日期选择器"""
        start_date = changelist.request.GET.get('start_date', '')
        end_date = changelist.request.GET.get('end_date', '')
        
        # 全部选项
        yield {
            'selected': self.value() is None and not start_date,
            'query_string': changelist.get_query_string(remove=[self.parameter_name, 'start_date', 'end_date']),
            'display': '全部时间',
        }
        
        # 快捷选项
        for lookup, title in self.lookup_choices:
            if lookup != 'custom':
                yield {
                    'selected': self.value() == lookup,
                    'query_string': changelist.get_query_string({self.parameter_name: lookup}, 
                                                              ['start_date', 'end_date']),
                    'display': title,
                }
        
        # 自定义日期范围选项
        yield {
            'selected': bool(start_date and end_date),
            'query_string': '',
            'display': '自定义时间范围',
            'custom_date': True,
            'start_date': start_date,
            'end_date': end_date,
            'current_params': dict(changelist.request.GET.lists()),
        }

    def expected_parameters(self):
        return [self.parameter_name, 'start_date', 'end_date']


class SummaryProjectFilter(SimpleListFilter):
    """汇总页面的项目筛选器"""
    title = '商机项目'
    parameter_name = 'project'

    def lookups(self, request, model_admin):
        return SalesOpportunity.PROJECT_CHOICES
    
    def queryset(self, request, queryset):
        if self.value():
            # 获取有该项目商机的医院调研
            hospital_surveys_with_project = SalesOpportunity.objects.filter(
                opportunity_project=self.value(),
                is_active=True
            ).values_list('hospital_survey_id', flat=True).distinct()
            
            return queryset.filter(hospital_survey_id__in=hospital_surveys_with_project)
        return queryset

 
class SalesOpportunitySummaryAdmin(GlobalAdmin):
    """销售商机汇总展示Admin - 按医院分组"""
    
    verbose_name = "医院商机汇总"
    verbose_name_plural = "医院商机汇总展示"
    
    list_display = [
        'hospital_name_display',
        'get_hospital_district',
        'get_hospital_class', 
        'get_hospital_manager',
        'get_opportunity_summary',
        'get_total_opportunities',
        'get_total_sample_volume',
        'get_latest_opportunity_date',
    ]
    
    list_filter = [
        'hospital_survey__hospital__district',
        'hospital_survey__hospital__hospitalclass', 
        SummaryHospitalManagerFilter,  # 修改这里
        SummaryDateRangeFilter,        # 修改这里
        MultiSelectMonthFilter,
        MultiSelectYearFilter,
        SummaryProjectFilter,          # 新增项目筛选
    ]

    search_fields = [
        'hospital_survey__hospital__hospitalname',
        'salesperson__chinesename'
    ]
    
    list_per_page = 50
    
    def get_queryset(self, request):
        """获取按医院分组的商机数据"""
        try:
            # 保存request以供筛选器使用
            self._current_request = request
            
            # 获取基础查询集
            base_queryset = super().get_queryset(request).filter(is_active=True)
            
            # # 应用权限过滤
            # if not request.user.is_superuser and request.user.username in ALLOWED_SALESMEN:
            #     base_queryset = base_queryset.filter(salesperson=request.user)
            
            # 按医院分组，每个医院只显示一条记录（最新的商机）
            from django.db.models import Max
            
            # 获取每个医院的最新商机ID
            latest_opportunity_ids = base_queryset.values('hospital_survey').annotate(
                latest_id=Max('id')
            ).values_list('latest_id', flat=True)
            
            # 只返回这些最新的商机记录
            result = base_queryset.filter(
                id__in=latest_opportunity_ids
            ).select_related(
                'hospital_survey__hospital', 
                'salesperson',
                'hospital_survey__qitian_manager'
            ).order_by('-createtime')
            
            return result
            
        except Exception as e:
            print(f"get_queryset error: {e}")
            import traceback
            traceback.print_exc()
            return super().get_queryset(request).filter(is_active=True).select_related(
                'hospital_survey__hospital', 
                'salesperson',
                'hospital_survey__qitian_manager'
            )

    def hospital_name_display(self, obj):
        """显示医院名称"""
        return obj.hospital_survey.hospital.hospitalname if obj.hospital_survey and obj.hospital_survey.hospital else '-'
    hospital_name_display.short_description = '医院名称'
    hospital_name_display.admin_order_field = 'hospital_survey__hospital__hospitalname'
    
    def get_hospital_district(self, obj):
        return obj.hospital_survey.hospital.district if obj.hospital_survey and obj.hospital_survey.hospital else '-'
    get_hospital_district.short_description = '区域'
    get_hospital_district.admin_order_field = 'hospital_survey__hospital__district'
    
    def get_hospital_class(self, obj):
        return obj.hospital_survey.hospital.hospitalclass if obj.hospital_survey and obj.hospital_survey.hospital else '-'
    get_hospital_class.short_description = '级别'
    get_hospital_class.admin_order_field = 'hospital_survey__hospital__hospitalclass'
    
    def get_hospital_manager(self, obj):
        return obj.salesperson.chinesename if obj.salesperson else '-'
    get_hospital_manager.short_description = '销售人员'
    get_hospital_manager.admin_order_field = 'salesperson__chinesename'

    def get_opportunity_summary(self, obj):
        """获取该医院的所有商机汇总信息，考虑当前筛选条件"""
        try:
            # 获取该医院的所有商机
            opportunities_query = SalesOpportunity.objects.filter(
                hospital_survey=obj.hospital_survey,
                is_active=True
            )
            
            # 应用当前的筛选条件
            request = getattr(self, '_current_request', None)
            if request:
                # 应用日期筛选
                opportunities_query = self._apply_date_filters(opportunities_query, request)
                
                # 应用项目筛选
                project = request.GET.get('project')
                if project:
                    opportunities_query = opportunities_query.filter(opportunity_project=project)
            
            opportunities = opportunities_query.order_by('-createtime')
            
            if not opportunities.exists():
                return '-'
            
            opportunity_summaries = []
            for opp in opportunities:
                summary = f"（{opp.get_opportunity_project_display()}-{opp.opportunity_model}-{opp.sample_volume or 0}-{opp.landing_time or '未定'}）"
                opportunity_summaries.append(summary)
            
            result = "/".join(opportunity_summaries)
            
            if len(result) > 200:
                return f"{result[:200]}..."
            return result
            
        except Exception as e:
            print(f"get_opportunity_summary error: {e}")
            return '-'
    
    get_opportunity_summary.short_description = '商机项目-型号-标本量-落地时间'
    
    def get_total_opportunities(self, obj):
        """获取该医院的商机总数，考虑筛选条件"""
        try:
            opportunities_query = SalesOpportunity.objects.filter(
                hospital_survey=obj.hospital_survey,
                is_active=True
            )
            
            # 应用筛选条件
            request = getattr(self, '_current_request', None)
            if request:
                opportunities_query = self._apply_date_filters(opportunities_query, request)
                
                project = request.GET.get('project')
                if project:
                    opportunities_query = opportunities_query.filter(opportunity_project=project)
            
            return opportunities_query.count()
        except Exception:
            return 0
    get_total_opportunities.short_description = '商机数量'
    
    def get_total_sample_volume(self, obj):
        """获取该医院的标本量总和，考虑筛选条件"""
        try:
            from django.db.models import Sum
            
            opportunities_query = SalesOpportunity.objects.filter(
                hospital_survey=obj.hospital_survey,
                is_active=True
            )
            
            # 应用筛选条件
            request = getattr(self, '_current_request', None)
            if request:
                opportunities_query = self._apply_date_filters(opportunities_query, request)
                
                project = request.GET.get('project')
                if project:
                    opportunities_query = opportunities_query.filter(opportunity_project=project)
            
            result = opportunities_query.aggregate(total=Sum('sample_volume'))
            return result['total'] or 0
        except Exception:
            return 0
    get_total_sample_volume.short_description = '标本量总和'
    
    def get_latest_opportunity_date(self, obj):
        """获取该医院最新商机创建时间，考虑筛选条件"""
        try:
            opportunities_query = SalesOpportunity.objects.filter(
                hospital_survey=obj.hospital_survey,
                is_active=True
            )
            
            # 应用筛选条件
            request = getattr(self, '_current_request', None)
            if request:
                opportunities_query = self._apply_date_filters(opportunities_query, request)
                
                project = request.GET.get('project')
                if project:
                    opportunities_query = opportunities_query.filter(opportunity_project=project)
            
            latest = opportunities_query.order_by('-createtime').first()
            
            if latest:
                from django.utils import timezone
                return timezone.localtime(latest.createtime).strftime('%Y-%m-%d')
            return '-'
        except Exception:
            return '-'
    get_latest_opportunity_date.short_description = '最新更新时间'
    
    def _apply_date_filters(self, queryset, request):
        """应用日期筛选条件的辅助方法"""
        # 处理自定义日期范围
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        
        if start_date and end_date:
            try:
                from datetime import datetime
                start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
                end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
                return queryset.filter(
                    createtime__date__gte=start_date_obj,
                    createtime__date__lte=end_date_obj
                )
            except ValueError:
                pass
        
        # 处理快捷日期选项
        date_range = request.GET.get('date_range')
        if date_range:
            from django.utils import timezone
            from datetime import timedelta
            import calendar
            
            today = timezone.now().date()
            
            if date_range == 'today':
                return queryset.filter(createtime__date=today)
            elif date_range == 'yesterday':
                yesterday = today - timedelta(days=1)
                return queryset.filter(createtime__date=yesterday)
            elif date_range == 'this_week':
                start_date_obj = today - timedelta(days=today.weekday())
                end_date_obj = start_date_obj + timedelta(days=6)
                return queryset.filter(
                    createtime__date__gte=start_date_obj,
                    createtime__date__lte=end_date_obj
                )
            elif date_range == 'last_week':
                start_of_this_week = today - timedelta(days=today.weekday())
                start_date_obj = start_of_this_week - timedelta(days=7)
                end_date_obj = start_of_this_week - timedelta(days=1)
                return queryset.filter(
                    createtime__date__gte=start_date_obj,
                    createtime__date__lte=end_date_obj
                )
            elif date_range == 'this_month':
                start_date_obj = today.replace(day=1)
                end_date_obj = today.replace(day=calendar.monthrange(today.year, today.month)[1])
                return queryset.filter(
                    createtime__date__gte=start_date_obj,
                    createtime__date__lte=end_date_obj
                )
            elif date_range == 'last_month':
                if today.month == 1:
                    last_month = today.replace(year=today.year-1, month=12, day=1)
                else:
                    last_month = today.replace(month=today.month-1, day=1)
                start_date_obj = last_month
                end_date_obj = last_month.replace(day=calendar.monthrange(last_month.year, last_month.month)[1])
                return queryset.filter(
                    createtime__date__gte=start_date_obj,
                    createtime__date__lte=end_date_obj
                )
            # 其他日期范围逻辑...
        
        return queryset
    
    def changelist_view(self, request, extra_context=None):
        """重写changelist_view以添加统计信息"""
        extra_context = extra_context or {}
        
        # 保存request以供其他方法使用
        self._current_request = request
        
        # 获取统计信息
        queryset = self.get_queryset(request)
        cl = self.get_changelist_instance(request)
        filtered_queryset = cl.get_queryset(request)
        
        # 计算统计数据
        total_hospitals = filtered_queryset.count()
        
        # 通过所有符合条件的医院调研，计算商机总数和标本总量
        hospital_survey_ids = filtered_queryset.values_list('hospital_survey_id', flat=True)
        
        # 构建商机查询，应用相同的筛选条件
        opportunities_query = SalesOpportunity.objects.filter(
            hospital_survey_id__in=hospital_survey_ids,
            is_active=True
        )
        
        # 应用筛选条件
        opportunities_query = self._apply_date_filters(opportunities_query, request)
        
        project = request.GET.get('project')
        if project:
            opportunities_query = opportunities_query.filter(opportunity_project=project)
        
        from django.db.models import Sum, Count
        stats = opportunities_query.aggregate(
            total_opportunities=Count('id'),
            total_sample_volume=Sum('sample_volume')
        )
        
        extra_context.update({
            'total_hospitals': total_hospitals,
            'total_opportunities': stats['total_opportunities'] or 0,
            'total_sample_volume': stats['total_sample_volume'] or 0,
            'is_salesperson': not request.user.is_superuser and request.user.username in ALLOWED_SALESMEN
        })
        
        return super().changelist_view(request, extra_context=extra_context)
    
    # 权限控制
    def has_add_permission(self, request):
        return False
    
    def has_change_permission(self, request, obj=None):
        return False
    
    def has_delete_permission(self, request, obj=None):
        return False


# 重新注册
admin.site.register(SalesOpportunitySummary, SalesOpportunitySummaryAdmin)

 
 

